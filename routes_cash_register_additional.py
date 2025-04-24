"""
Rutas adicionales para el módulo de Arqueos de Caja.

Este archivo contiene las rutas para la gestión centralizada de arqueos, visualización,
edición y eliminación de arqueos.
"""

import logging
from datetime import datetime, date, timedelta
import calendar
import os
import io
import csv
from decimal import Decimal

# Imports de Flask y extensiones
from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, 
    jsonify, current_app, abort, session, send_file
)
from flask_login import login_required, current_user
from werkzeug.security import safe_join

# Imports de modelos, formularios y utilidades
from app import db
from models import Company, Employee, User
from models_cash_register import CashRegister, CashRegisterToken, CashRegisterSummary
from forms_cash_register import CashRegisterForm, CashRegisterSearchForm
from utils_cash_register import (
    calculate_weekly_summary, calculate_staff_cost, calculate_monthly_revenue,
    calculate_yearly_revenue, format_currency, format_percentage,
    get_current_week_number, get_week_dates, get_week_number, get_date_range
)

# Configuración del logger
logger = logging.getLogger(__name__)


# Funciones auxiliares 
def check_company_access(company):
    """Verifica que el usuario actual tenga acceso a la empresa."""
    if not current_user.is_admin() and company.id not in [c.id for c in current_user.companies]:
        logger.warning(f"Usuario {current_user.id} intentó acceder a empresa {company.id} sin permisos")
        abort(403)
    return True


def get_user_companies(user):
    """Retorna las empresas a las que tiene acceso el usuario."""
    if user.is_admin():
        return Company.query.all()
    return user.companies


# Definición de las rutas adicionales para gestión centralizada de arqueos
def register_routes(cash_register_bp):
    """
    Registra las rutas adicionales en el blueprint de arqueos de caja.
    
    Args:
        cash_register_bp: Blueprint de arqueos de caja
    """
    # Verificar rutas existentes para evitar duplicados
    existing_endpoints = set()
    for rule in cash_register_bp.deferred_functions:
        if hasattr(rule, 'endpoint'):
            existing_endpoints.add(rule.endpoint)
            
    logger.info(f"Rutas existentes: {existing_endpoints}")
    
    @cash_register_bp.route('/manage/<int:company_id>')
    @login_required
    def manage_registers(company_id):
        """Vista para gestionar (ver, editar, eliminar) arqueos de una empresa."""
        logger.info(f"Acceso a gestión centralizada de arqueos para empresa {company_id}")
        
        # Verificar acceso a la empresa
        company = Company.query.get_or_404(company_id)
        check_company_access(company)
        
        # Obtener parámetros de filtrado
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', 0, type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        limit = request.args.get('limit', 20, type=int)
        
        # Iniciar consulta base
        query = CashRegister.query.filter_by(company_id=company_id)
        
        # Aplicar filtros
        if start_date and end_date:
            # Filtrar por rango de fechas específico
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(
                CashRegister.date >= start_date,
                CashRegister.date <= end_date
            )
        elif month > 0:
            # Filtrar por mes y año específicos
            start_date, end_date = get_date_range(year, month)
            query = query.filter(
                CashRegister.date >= start_date,
                CashRegister.date <= end_date
            )
        elif year:
            # Filtrar solo por año
            start_date = date(year, 1, 1)
            end_date = date(year, 12, 31)
            query = query.filter(
                CashRegister.date >= start_date,
                CashRegister.date <= end_date
            )
            
        # Aplicar ordenamiento (más recientes primero)
        query = query.order_by(CashRegister.date.desc())
        
        # Aplicar límite si es necesario
        if limit > 0:
            registers = query.limit(limit).all()
        else:
            registers = query.all()
            
        # Calcular totales
        totals = {
            'cash': sum(r.cash_amount for r in registers),
            'card': sum(r.card_amount for r in registers),
            'delivery_cash': sum(r.delivery_cash_amount for r in registers),
            'delivery_online': sum(r.delivery_online_amount for r in registers),
            'check': sum(r.check_amount for r in registers),
            'expenses': sum(r.expenses_amount for r in registers),
            'total': sum(r.total_amount for r in registers)
        }
        
        return render_template(
            'cash_register/manage_registers.html',
            company=company,
            registers=registers,
            totals=totals,
            current_year=datetime.now().year
        )


    @cash_register_bp.route('/view/<int:register_id>')
    @login_required
    def view_register(register_id):
        """Vista de detalles de un arqueo específico."""
        logger.info(f"Visualizando detalles del arqueo {register_id}")
        
        # Obtener el arqueo y verificar acceso
        register = CashRegister.query.get_or_404(register_id)
        company = Company.query.get_or_404(register.company_id)
        check_company_access(company)
        
        # Obtener la URL de retorno
        back_url = request.referrer
        
        return render_template(
            'cash_register/view_register.html',
            register=register,
            company=company,
            back_url=back_url
        )


    # No se registra la ruta /delete/<int:register_id> porque ya existe en routes_cash_register.py
    # En su lugar, usamos el endpoint existente
        
        
    @cash_register_bp.route('/export/<int:company_id>')
    @login_required
    def export_registers(company_id):
        """Exporta los arqueos de una empresa a CSV o Excel."""
        logger.info(f"Exportando arqueos de empresa {company_id}")
        
        # Verificar acceso a la empresa
        company = Company.query.get_or_404(company_id)
        check_company_access(company)
        
        # Obtener formato de exportación
        export_format = request.args.get('format', 'csv')
        
        # Obtener parámetros de filtrado
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', 0, type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Iniciar consulta base
        query = CashRegister.query.filter_by(company_id=company_id)
        
        # Aplicar filtros (código similar a manage_registers)
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(
                CashRegister.date >= start_date,
                CashRegister.date <= end_date
            )
        elif month > 0:
            start_date, end_date = get_date_range(year, month)
            query = query.filter(
                CashRegister.date >= start_date,
                CashRegister.date <= end_date
            )
        elif year:
            start_date = date(year, 1, 1)
            end_date = date(year, 12, 31)
            query = query.filter(
                CashRegister.date >= start_date,
                CashRegister.date <= end_date
            )
            
        # Obtener registros ordenados por fecha
        registers = query.order_by(CashRegister.date.desc()).all()
        
        if export_format == 'csv':
            # Exportar a CSV
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Cabecera
            writer.writerow([
                'Fecha', 'Efectivo', 'Tarjeta', 'Delivery Efectivo', 
                'Delivery Online', 'Cheque', 'Gastos', 'Total',
                'Notas Gastos', 'Notas', 'Empleado'
            ])
            
            # Datos
            for reg in registers:
                writer.writerow([
                    reg.date.strftime('%d/%m/%Y'),
                    f"{reg.cash_amount:.2f}",
                    f"{reg.card_amount:.2f}",
                    f"{reg.delivery_cash_amount:.2f}",
                    f"{reg.delivery_online_amount:.2f}",
                    f"{reg.check_amount:.2f}",
                    f"{reg.expenses_amount:.2f}",
                    f"{reg.total_amount:.2f}",
                    reg.expenses_notes,
                    reg.notes,
                    reg.employee_name
                ])
            
            # Configurar respuesta
            output.seek(0)
            filename = f"arqueos_{company.name}_{datetime.now().strftime('%Y%m%d')}.csv"
            return send_file(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                mimetype='text/csv',
                as_attachment=True,
                download_name=filename
            )
            
        elif export_format == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
                
                # Crear libro de Excel
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Arqueos de Caja"
                
                # Estilo para encabezados
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="0366D6", end_color="0366D6", fill_type="solid")
                
                # Cabecera
                headers = [
                    'Fecha', 'Efectivo', 'Tarjeta', 'Delivery Efectivo', 
                    'Delivery Online', 'Cheque', 'Gastos', 'Total',
                    'Notas Gastos', 'Notas', 'Empleado'
                ]
                
                # Aplicar estilos a la cabecera
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Datos
                for row, reg in enumerate(registers, 2):
                    ws.cell(row=row, column=1, value=reg.date.strftime('%d/%m/%Y'))
                    ws.cell(row=row, column=2, value=reg.cash_amount)
                    ws.cell(row=row, column=3, value=reg.card_amount)
                    ws.cell(row=row, column=4, value=reg.delivery_cash_amount)
                    ws.cell(row=row, column=5, value=reg.delivery_online_amount)
                    ws.cell(row=row, column=6, value=reg.check_amount)
                    ws.cell(row=row, column=7, value=reg.expenses_amount)
                    ws.cell(row=row, column=8, value=reg.total_amount)
                    ws.cell(row=row, column=9, value=reg.expenses_notes)
                    ws.cell(row=row, column=10, value=reg.notes)
                    ws.cell(row=row, column=11, value=reg.employee_name)
                
                # Ajustar ancho de columnas
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                # Guardar a bytes
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # Configurar respuesta
                filename = f"arqueos_{company.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
            except ImportError:
                flash('No se pudo exportar a Excel. Librería openpyxl no disponible.', 'danger')
                return redirect(url_for('cash_register.manage_registers', company_id=company_id))
        
        else:
            flash('Formato de exportación no soportado.', 'danger')
            return redirect(url_for('cash_register.manage_registers', company_id=company_id))


    @cash_register_bp.route('/print/<int:register_id>')
    @login_required
    def print_register(register_id):
        """Genera una vista imprimible de un arqueo."""
        logger.info(f"Generando vista de impresión para arqueo {register_id}")
        
        # Obtener el arqueo y verificar acceso
        register = CashRegister.query.get_or_404(register_id)
        company = Company.query.get_or_404(register.company_id)
        check_company_access(company)
        
        # Renderizar plantilla de impresión
        return render_template(
            'cash_register/print_register.html',
            register=register,
            company=company,
            format_currency=format_currency
        )
    
    return cash_register_bp