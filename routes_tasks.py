from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, session, current_app, send_file, Response
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime, date, timedelta
import os
import io
import socket
import openpyxl
from fpdf import FPDF
from werkzeug.utils import secure_filename
from wtforms.validators import Optional

from app import db
from models import User, Company, Employee
from models_tasks import (Location, LocalUser, Task, TaskSchedule, TaskCompletion, TaskPriority, 
                         TaskFrequency, TaskStatus, WeekDay, TaskGroup, TaskWeekday, TaskMonthDay,
                         Product, ProductConservation, ProductLabel, ConservationType, LabelTemplate,
                         NetworkPrinter, TaskInstance)
from forms_tasks import (LocationForm, LocalUserForm, TaskForm, DailyScheduleForm, WeeklyScheduleForm, 
                        MonthlyScheduleForm, BiweeklyScheduleForm, TaskCompletionForm, 
                        LocalUserPinForm, SearchForm, TaskGroupForm, CustomWeekdaysForm, PortalLoginForm,
                        ProductForm, ProductConservationForm, GenerateLabelForm, LabelEditorForm,
                        NetworkPrinterForm)
from utils import log_activity, can_manage_company, save_file
from utils_tasks import create_default_local_user, regenerate_portal_password, count_available_employees, sync_employees_to_local_users
from weekly_tasks_reset_service import reset_weekly_tasks, process_custom_tasks_for_week

# Crear el Blueprint para las tareas
tasks_bp = Blueprint('tasks', __name__)

# Decorador para proteger rutas y asegurar que el usuario es admin o gerente
def manager_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        
        if not (current_user.is_admin() or current_user.is_gerente()):
            flash('No tienes permiso para acceder a esta página.', 'danger')
            return redirect(url_for('index'))
            
        return f(*args, **kwargs)
    return decorated_function

# Decorador para proteger rutas de usuario local
def local_user_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'local_user_id' not in session:
            return redirect(url_for('tasks.local_login'))
            
        return f(*args, **kwargs)
    return decorated_function

# Rutas para el dashboard principal de tareas
@tasks_bp.route('/')
@login_required
def index():
    """Dashboard principal del módulo de tareas"""
    # Determinar qué información mostrar según el rol del usuario
    if current_user.is_admin():
        # Los administradores ven todos los locales
        locations = Location.query.all()
    elif current_user.is_gerente():
        # Los gerentes ven solo los locales de sus empresas
        company_ids = [c.id for c in current_user.companies]
        locations = Location.query.filter(Location.company_id.in_(company_ids)).all()
    else:
        # Otros usuarios no tendrían acceso, pero por si acaso
        locations = []
    
    # Si hay locales, redirigir al primero
    if locations:
        first_location = locations[0]
        return redirect(url_for('tasks.view_location', id=first_location.id))
        
    # Si no hay locales, mostrar la página de dashboard con el botón para crear el primer local
    return render_template('tasks/dashboard.html', locations=locations)

# Rutas para gestión de locales
@tasks_bp.route('/locations')
@login_required
@manager_required
def list_locations():
    """Lista de locales disponibles"""
    if current_user.is_admin():
        locations = Location.query.all()
    elif current_user.is_gerente() and current_user.company_id:
        locations = Location.query.filter_by(company_id=current_user.company_id).all()
    else:
        locations = []
    
    return render_template('tasks/location_list.html',
                          title='Gestión de Locales',
                          locations=locations)

@tasks_bp.route('/locations/create', methods=['GET', 'POST'])
@login_required
@manager_required
def create_location():
    """Crear un nuevo local"""
    form = LocationForm()
    
    # Cargar empresas disponibles según el rol
    if current_user.is_admin():
        form.company_id.choices = [(c.id, c.name) for c in Company.query.all()]
    elif current_user.is_gerente() and current_user.company_id:
        company = Company.query.get(current_user.company_id)
        if company:
            form.company_id.choices = [(company.id, company.name)]
        else:
            form.company_id.choices = []
    else:
        form.company_id.choices = []
    
    if form.validate_on_submit():
        location = Location(
            name=form.name.data,
            address=form.address.data,
            city=form.city.data,
            postal_code=form.postal_code.data,
            description=form.description.data,
            company_id=form.company_id.data,
            is_active=form.is_active.data
        )
        
        # Si se proporcionó contraseña, establecerla de forma segura
        if form.portal_password.data:
            location.set_portal_password(form.portal_password.data)
        
        db.session.add(location)
        db.session.commit()
        
        log_activity(f'Local creado: {location.name}')
        flash(f'Local "{location.name}" creado correctamente.', 'success')
        return redirect(url_for('tasks.list_locations'))
    
    return render_template('tasks/location_form.html',
                          title='Crear Nuevo Local',
                          form=form)

@tasks_bp.route('/locations/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@manager_required
def edit_location(id):
    """Editar un local existente"""
    location = Location.query.get_or_404(id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para editar este local.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    form = LocationForm(obj=location)
    
    # No cargar contraseña del portal, ya que la tenemos como hash
    form.portal_password.data = None
    form.confirm_portal_password.data = None
    
    # Cargar empresas disponibles según el rol
    if current_user.is_admin():
        form.company_id.choices = [(c.id, c.name) for c in Company.query.all()]
    elif current_user.is_gerente() and current_user.company_id:
        company = Company.query.get(current_user.company_id)
        if company:
            form.company_id.choices = [(company.id, company.name)]
        else:
            form.company_id.choices = []
    else:
        form.company_id.choices = []
    
    if form.validate_on_submit():
        location.name = form.name.data
        location.address = form.address.data
        location.city = form.city.data
        location.postal_code = form.postal_code.data
        location.description = form.description.data
        location.company_id = form.company_id.data
        location.is_active = form.is_active.data
        # Ya no se permite personalizar el nombre de usuario
        
        # Si se proporcionó una nueva contraseña, actualizarla
        if form.portal_password.data:
            location.set_portal_password(form.portal_password.data)
        
        db.session.commit()
        
        log_activity(f'Local actualizado: {location.name}')
        flash(f'Local "{location.name}" actualizado correctamente.', 'success')
        return redirect(url_for('tasks.list_locations'))
    
    return render_template('tasks/location_form.html',
                          title=f'Editar Local: {location.name}',
                          form=form,
                          location=location)

@tasks_bp.route('/locations/delete/<int:id>', methods=['POST'])
@login_required
@manager_required
def delete_location(id):
    """Eliminar un local"""
    location = Location.query.get_or_404(id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para eliminar este local.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    # Verificar si tiene tareas o usuarios asociados
    if location.tasks or location.local_users:
        flash('No se puede eliminar este local porque tiene tareas o usuarios asociados.', 'warning')
        return redirect(url_for('tasks.list_locations'))
    
    name = location.name
    db.session.delete(location)
    db.session.commit()
    
    log_activity(f'Local eliminado: {name}')
    flash(f'Local "{name}" eliminado correctamente.', 'success')
    return redirect(url_for('tasks.list_locations'))

@tasks_bp.route('/locations/<int:id>/groups')
@login_required
@manager_required
def list_task_groups(id):
    """Lista de grupos de tareas para un local"""
    location = Location.query.get_or_404(id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para ver los grupos de este local.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    # Obtener todos los grupos para este local
    groups = TaskGroup.query.filter_by(location_id=id).all()
    
    return render_template('tasks/task_group_list.html',
                          title=f'Grupos de Tareas - {location.name}',
                          location=location,
                          groups=groups)

@tasks_bp.route('/locations/<int:location_id>/groups/create', methods=['GET', 'POST'])
@login_required
@manager_required
def create_task_group(location_id):
    """Crear un nuevo grupo de tareas"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para crear grupos en este local.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    form = TaskGroupForm()
    form.location_id.choices = [(location.id, location.name)]
    form.location_id.data = location.id
    
    if form.validate_on_submit():
        group = TaskGroup(
            name=form.name.data,
            description=form.description.data,
            color=form.color.data,
            location_id=location.id
        )
        
        db.session.add(group)
        db.session.commit()
        
        log_activity(f'Grupo de tareas creado: {group.name} para local {location.name}')
        flash(f'Grupo "{group.name}" creado correctamente.', 'success')
        return redirect(url_for('tasks.list_task_groups', id=location.id))
    
    return render_template('tasks/task_group_form.html',
                          title='Crear Nuevo Grupo de Tareas',
                          form=form,
                          location=location)

@tasks_bp.route('/task-groups/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@manager_required
def edit_task_group(id):
    """Editar un grupo de tareas existente"""
    group = TaskGroup.query.get_or_404(id)
    location = group.location
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para editar este grupo.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    form = TaskGroupForm(obj=group)
    form.location_id.choices = [(location.id, location.name)]
    
    if form.validate_on_submit():
        group.name = form.name.data
        group.description = form.description.data
        group.color = form.color.data
        
        db.session.commit()
        
        log_activity(f'Grupo de tareas actualizado: {group.name}')
        flash(f'Grupo "{group.name}" actualizado correctamente.', 'success')
        return redirect(url_for('tasks.list_task_groups', id=location.id))
    
    return render_template('tasks/task_group_form.html',
                          title=f'Editar Grupo: {group.name}',
                          form=form,
                          group=group,
                          location=location)

@tasks_bp.route('/task-groups/<int:id>/delete', methods=['POST'])
@login_required
@manager_required
def delete_task_group(id):
    """Eliminar un grupo de tareas"""
    group = TaskGroup.query.get_or_404(id)
    location = group.location
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para eliminar este grupo.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    # Verificar si hay tareas asignadas al grupo
    tasks = Task.query.filter_by(group_id=id).all()
    if tasks:
        # Actualizar las tareas para desasociarlas del grupo
        for task in tasks:
            task.group_id = None
        db.session.commit()
    
    name = group.name
    location_id = location.id
    db.session.delete(group)
    db.session.commit()
    
    log_activity(f'Grupo de tareas eliminado: {name}')
    flash(f'Grupo "{name}" eliminado correctamente.', 'success')
    return redirect(url_for('tasks.list_task_groups', id=location_id))

@tasks_bp.route('/locations/<int:id>')
@login_required
def view_location(id):
    """Ver detalles de un local"""
    location = Location.query.get_or_404(id)
    
    # Verificar permisos (admin, gerente de la empresa, o empleado de la empresa)
    if not current_user.is_admin() and not (current_user.company_id == location.company_id):
        flash('No tienes permiso para ver este local.', 'danger')
        return redirect(url_for('tasks.index'))
    
    # Obtener tareas para hoy
    today = date.today()
    active_tasks = []
    pending_tasks = []
    
    # Obtener todas las tareas del local
    location_tasks = Task.query.filter_by(location_id=location.id).all()
    
    # Clasificar tareas según su estado
    completed_tasks = []
    for task in location_tasks:
        if task.status == TaskStatus.COMPLETADA:
            completed_tasks.append(task)
        elif task.status == TaskStatus.PENDIENTE:
            pending_tasks.append(task)
            # Si además está programada para hoy, es una tarea activa
            if task.is_due_today():
                active_tasks.append(task)
    
    # Obtener usuarios locales
    local_users = LocalUser.query.filter_by(location_id=location.id).all()
    
    # Obtener tareas recientes completadas
    recent_completions = (db.session.query(TaskCompletion, Task, LocalUser)
                         .join(Task, TaskCompletion.task_id == Task.id)
                         .join(LocalUser, TaskCompletion.local_user_id == LocalUser.id)
                         .filter(Task.location_id == location.id)
                         .order_by(TaskCompletion.completion_date.desc())
                         .limit(10)
                         .all())
    
    # Generar datos para el gráfico de completados mensuales
    last_6_months = []
    monthly_counts = []
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    for i in range(5, -1, -1):
        month = (current_month - i) % 12
        if month == 0:
            month = 12
        year = current_year if month <= current_month else current_year - 1
        
        month_start = datetime(year, month, 1)
        if month == 12:
            next_month_start = datetime(year + 1, 1, 1)
        else:
            next_month_start = datetime(year, month + 1, 1)
            
        # Obtener completados de este mes
        month_count = TaskCompletion.query\
            .join(Task, TaskCompletion.task_id == Task.id)\
            .filter(Task.location_id == location.id)\
            .filter(TaskCompletion.completion_date >= month_start)\
            .filter(TaskCompletion.completion_date < next_month_start)\
            .count()
            
        # Formato del mes en español
        month_names = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        month_label = f"{month_names[month-1]} {year}"
        
        last_6_months.append(month_label)
        monthly_counts.append(month_count)
    
    # Obtener el enlace directo al portal
    portal_url = url_for('tasks.local_portal', location_id=location.id, _external=True)
    
    return render_template('tasks/location_detail.html',
                          title=f'Local: {location.name}',
                          location=location,
                          active_tasks=active_tasks,
                          pending_tasks=pending_tasks,
                          completed_tasks=completed_tasks,
                          local_users=local_users,
                          recent_completions=recent_completions,
                          monthly_labels=last_6_months,
                          monthly_counts=monthly_counts,
                          portal_url=portal_url)

# La función sync_employees_to_local_users ha sido movida a utils_tasks.py para mejor organización

# Rutas para gestión de usuarios locales
@tasks_bp.route('/locations/<int:location_id>/users')
@login_required
@manager_required
def list_local_users(location_id):
    """Lista de usuarios de un local"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para ver los usuarios de este local.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    users = LocalUser.query.filter_by(location_id=location_id).all()
    
    # Contar empleados disponibles para importar utilizando la función especializada
    # que excluye los que ya están importados como usuarios locales
    available_employees_count = count_available_employees(location_id)
    
    # Contar usuarios ya importados
    imported_users_count = LocalUser.query.filter_by(
        location_id=location_id,
        imported=True
    ).count()
    
    return render_template('tasks/local_user_list.html',
                          title=f'Usuarios del Local: {location.name}',
                          location=location,
                          users=users,
                          available_employees_count=available_employees_count,
                          imported_users_count=imported_users_count)

@tasks_bp.route('/locations/<int:location_id>/users/sync-employees', methods=['POST'])
@login_required
@manager_required
def sync_employees(location_id):
    """Sincronizar empleados de la empresa como usuarios locales"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para sincronizar empleados en este local.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    # Realizar la sincronización
    try:
        total_employees, created, updated = sync_employees_to_local_users(location_id)
        
        if created > 0 or updated > 0:
            flash(f'Sincronización completada: {created} empleados nuevos importados, {updated} empleados actualizados.', 'success')
            log_activity(f'Empleados sincronizados en {location.name}: {created} nuevos, {updated} actualizados')
        else:
            flash('No se encontraron cambios para sincronizar.', 'info')
            
    except Exception as e:
        flash(f'Error al sincronizar empleados: {str(e)}', 'danger')
        log_activity(f'Error al sincronizar empleados en {location.name}: {str(e)}', level='error')
    
    return redirect(url_for('tasks.list_local_users', location_id=location_id))

@tasks_bp.route('/locations/<int:location_id>/users/create', methods=['GET', 'POST'])
@login_required
@manager_required
def create_local_user(location_id):
    """Crear un nuevo usuario local"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para crear usuarios en este local.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    form = LocalUserForm()
    
    if form.validate_on_submit():
        # Generar nombre de usuario automáticamente (nombre + apellido + timestamp)
        timestamp = datetime.now().strftime('%y%m%d%H%M%S')
        username = f"{form.name.data.lower()}_{form.last_name.data.lower()}_{timestamp}"
        
        user = LocalUser(
            name=form.name.data,
            last_name=form.last_name.data,
            username=username,
            location_id=location_id,
            is_active=form.is_active.data
        )
        
        # Establecer PIN de acceso
        user.set_pin(form.pin.data)
        
        # Guardar foto si se proporciona
        if form.photo.data:
            filename = secure_filename(form.photo.data.filename)
            # Generar nombre único para evitar colisiones
            unique_filename = f"user_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
            photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'local_users', unique_filename)
            
            # Asegurar que existe el directorio
            os.makedirs(os.path.dirname(photo_path), exist_ok=True)
            
            # Guardar archivo
            form.photo.data.save(photo_path)
            user.photo_path = os.path.join('local_users', unique_filename)
        
        db.session.add(user)
        db.session.commit()
        
        log_activity(f'Usuario local creado: {user.name} en {location.name}')
        flash(f'Usuario "{user.name}" creado correctamente.', 'success')
        return redirect(url_for('tasks.list_local_users', location_id=location_id))
    
    return render_template('tasks/local_user_form.html',
                          title=f'Crear Usuario en {location.name}',
                          form=form,
                          location=location)

@tasks_bp.route('/locations/<int:location_id>/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@manager_required
def edit_local_user(location_id, id):
    """Editar un usuario local existente"""
    location = Location.query.get_or_404(location_id)
    user = LocalUser.query.get_or_404(id)
    
    # Verificar que el usuario pertenece al local
    if user.location_id != location_id:
        flash('El usuario no pertenece a este local.', 'danger')
        return redirect(url_for('tasks.list_local_users', location_id=location_id))
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para editar usuarios en este local.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    # Crear formulario, PIN opcional para editarlo
    form = LocalUserForm(obj=user)
    form.pin.validators = [Optional()]  # Hacer el PIN opcional
    
    if form.validate_on_submit():
        user.name = form.name.data
        user.last_name = form.last_name.data
        user.is_active = form.is_active.data
        
        # Actualizar PIN solo si se proporciona
        if form.pin.data:
            user.set_pin(form.pin.data)
        
        # Actualizar foto si se proporciona
        if form.photo.data:
            # Eliminar foto anterior si existe
            if user.photo_path:
                old_photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], user.photo_path)
                if os.path.exists(old_photo_path):
                    os.remove(old_photo_path)
            
            filename = secure_filename(form.photo.data.filename)
            # Generar nombre único para evitar colisiones
            unique_filename = f"user_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
            photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'local_users', unique_filename)
            
            # Asegurar que existe el directorio
            os.makedirs(os.path.dirname(photo_path), exist_ok=True)
            
            # Guardar archivo
            form.photo.data.save(photo_path)
            user.photo_path = os.path.join('local_users', unique_filename)
        
        db.session.commit()
        
        log_activity(f'Usuario local actualizado: {user.name} en {location.name}')
        flash(f'Usuario "{user.name}" actualizado correctamente.', 'success')
        return redirect(url_for('tasks.list_local_users', location_id=location_id))
    
    # Eliminar valores del campo PIN para seguridad
    form.pin.data = ''
    
    return render_template('tasks/local_user_form.html',
                          title=f'Editar Usuario: {user.name}',
                          form=form,
                          location=location,
                          user=user)

@tasks_bp.route('/locations/<int:location_id>/users/delete/<int:id>', methods=['POST'])
@login_required
@manager_required
def delete_local_user(location_id, id):
    """Eliminar un usuario local"""
    location = Location.query.get_or_404(location_id)
    user = LocalUser.query.get_or_404(id)
    
    # Verificar que el usuario pertenece al local
    if user.location_id != location_id:
        flash('El usuario no pertenece a este local.', 'danger')
        return redirect(url_for('tasks.list_local_users', location_id=location_id))
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para eliminar usuarios en este local.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    # Verificar si tiene tareas completadas asociadas
    if user.completed_tasks:
        flash('No se puede eliminar este usuario porque tiene tareas completadas asociadas.', 'warning')
        return redirect(url_for('tasks.list_local_users', location_id=location_id))
    
    # Eliminar foto si existe
    if user.photo_path:
        photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], user.photo_path)
        if os.path.exists(photo_path):
            os.remove(photo_path)
    
    name = user.name
    db.session.delete(user)
    db.session.commit()
    
    log_activity(f'Usuario local eliminado: {name} de {location.name}')
    flash(f'Usuario "{name}" eliminado correctamente.', 'success')
    return redirect(url_for('tasks.list_local_users', location_id=location_id))

# Rutas para gestión de tareas
@tasks_bp.route('/locations/<int:location_id>/tasks')
@login_required
def list_tasks(location_id):
    """Lista de tareas de un local"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin, gerente o empleado de la empresa)
    if not current_user.is_admin() and not (current_user.company_id == location.company_id):
        flash('No tienes permiso para ver las tareas de este local.', 'danger')
        return redirect(url_for('tasks.index'))
    
    # Filtrar por estado si se especifica
    status_filter = request.args.get('status')
    if status_filter and status_filter in [s.value for s in TaskStatus]:
        tasks = Task.query.filter_by(location_id=location_id, status=TaskStatus(status_filter)).all()
    else:
        tasks = Task.query.filter_by(location_id=location_id).all()
    
    # Agrupar tareas por estado
    tasks_by_status = {
        'pendiente': [],
        'completada': [],
        'vencida': [],
        'cancelada': []
    }
    
    for task in tasks:
        if task.status:
            tasks_by_status[task.status.value].append(task)
    
    return render_template('tasks/task_list.html',
                          title=f'Tareas del Local: {location.name}',
                          location=location,
                          tasks=tasks,
                          tasks_by_status=tasks_by_status)

@tasks_bp.route('/locations/<int:location_id>/tasks/create', methods=['GET', 'POST'])
@login_required
@manager_required
def create_task(location_id):
    """Crear una nueva tarea"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para crear tareas en este local.', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    form = TaskForm()
    form.location_id.choices = [(location.id, location.name)]
    
    # Cargar grupos de tareas disponibles para este local
    groups = TaskGroup.query.filter_by(location_id=location_id).all()
    group_choices = [(0, 'Ningún grupo')] + [(g.id, g.name) for g in groups]
    form.group_id.choices = group_choices
    
    if form.validate_on_submit():
        task = Task(
            title=form.title.data,
            description=form.description.data,
            priority=TaskPriority(form.priority.data),
            frequency=TaskFrequency(form.frequency.data),
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            location_id=form.location_id.data,
            created_by_id=current_user.id,
            status=TaskStatus.PENDIENTE
        )
        
        # Asignar grupo de tareas si se seleccionó uno
        if form.group_id.data and form.group_id.data != 0:
            task.group_id = form.group_id.data
        
        db.session.add(task)
        db.session.commit()
        
        # Para tareas personalizadas, guardar los días seleccionados
        if task.frequency == TaskFrequency.PERSONALIZADA:
            # Crear registros de días seleccionados
            if form.monday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.LUNES))
            if form.tuesday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.MARTES))
            if form.wednesday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.MIERCOLES))
            if form.thursday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.JUEVES))
            if form.friday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.VIERNES))
            if form.saturday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.SABADO))
            if form.sunday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.DOMINGO))
            
            db.session.commit()
        
        # Redirigir a la página de programación según la frecuencia
        log_activity(f'Tarea creada: {task.title} en {location.name}')
        flash(f'Tarea "{task.title}" creada correctamente. Ahora debes configurar su programación.', 'success')
        
        if task.frequency == TaskFrequency.DIARIA:
            return redirect(url_for('tasks.configure_daily_schedule', task_id=task.id))
        elif task.frequency == TaskFrequency.SEMANAL:
            return redirect(url_for('tasks.configure_weekly_schedule', task_id=task.id))
        elif task.frequency == TaskFrequency.MENSUAL:
            return redirect(url_for('tasks.configure_monthly_schedule', task_id=task.id))
        elif task.frequency == TaskFrequency.QUINCENAL:
            return redirect(url_for('tasks.configure_biweekly_schedule', task_id=task.id))
        else:
            return redirect(url_for('tasks.list_tasks', location_id=location_id))
    
    return render_template('tasks/task_form.html',
                          title=f'Crear Tarea en {location.name}',
                          form=form,
                          location=location,
                          location_id=location_id)

@tasks_bp.route('/tasks/<int:task_id>/daily-schedule', methods=['GET', 'POST'])
@login_required
@manager_required
def configure_daily_schedule(task_id):
    """Configurar frecuencia diaria para una tarea"""
    task = Task.query.get_or_404(task_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != task.location.company_id):
        flash('No tienes permiso para configurar esta tarea.', 'danger')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    # Verificar que sea una tarea diaria
    if task.frequency != TaskFrequency.DIARIA:
        flash('Esta tarea no es de frecuencia diaria.', 'warning')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    form = DailyScheduleForm()
    
    if form.validate_on_submit():
        log_activity(f'Frecuencia diaria configurada para tarea: {task.title}')
        flash('Configuración guardada correctamente.', 'success')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    return render_template('tasks/schedule_form.html',
                          title=f'Configurar Tarea Diaria: {task.title}',
                          form=form,
                          task=task,
                          schedule_type='diario')

@tasks_bp.route('/tasks/<int:task_id>/weekly-schedule', methods=['GET', 'POST'])
@login_required
@manager_required
def configure_weekly_schedule(task_id):
    """Configurar frecuencia semanal para una tarea"""
    task = Task.query.get_or_404(task_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != task.location.company_id):
        flash('No tienes permiso para configurar esta tarea.', 'danger')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    # Verificar que sea una tarea semanal
    if task.frequency != TaskFrequency.SEMANAL:
        flash('Esta tarea no es de frecuencia semanal.', 'warning')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    # Inicialización del formulario
    form = WeeklyScheduleForm()
    
    # Por defecto, todos los días seleccionados para tareas semanales
    form.monday.data = True
    form.tuesday.data = True
    form.wednesday.data = True
    form.thursday.data = True
    form.friday.data = True
    form.saturday.data = True
    form.sunday.data = True
    
    if form.validate_on_submit():
        log_activity(f'Frecuencia semanal configurada para tarea: {task.title}')
        flash('Configuración guardada correctamente.', 'success')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    return render_template('tasks/schedule_form.html',
                          title=f'Configurar Tarea Semanal: {task.title}',
                          form=form,
                          task=task,
                          schedule_type='semanal')

@tasks_bp.route('/tasks/<int:task_id>/monthly-schedule', methods=['GET', 'POST'])
@login_required
@manager_required
def configure_monthly_schedule(task_id):
    """Configurar días del mes para una tarea mensual"""
    task = Task.query.get_or_404(task_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != task.location.company_id):
        flash('No tienes permiso para configurar esta tarea.', 'danger')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    # Verificar que sea una tarea mensual
    if task.frequency != TaskFrequency.MENSUAL:
        flash('Esta tarea no es de frecuencia mensual.', 'warning')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    # Buscar días configurados con el nuevo modelo
    existing_days = [str(day.day_of_month) for day in task.monthdays]
    
    # Inicializar formulario
    form = MonthlyScheduleForm()
    
    # Precargar los días seleccionados del nuevo modelo
    # Siempre inicializar form.selected_days.data como lista aunque esté vacía
    form.selected_days.data = existing_days if existing_days else []
    
    if form.validate_on_submit():
        # Manejar el nuevo modelo (TaskMonthDay) - Primero eliminar días existentes
        for monthday in task.monthdays:
            db.session.delete(monthday)
        
        # Agregar los nuevos días seleccionados
        if form.selected_days.data:
            for day_str in form.selected_days.data:
                day_num = int(day_str)
                monthday = TaskMonthDay(task_id=task.id, day_of_month=day_num)
                db.session.add(monthday)
        
        # Guardar todos los cambios
        db.session.commit()
        
        log_activity(f'Días del mes configurados para tarea: {task.title}')
        flash('Configuración guardada correctamente.', 'success')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    return render_template('tasks/schedule_form.html',
                          title=f'Configurar Días del Mes: {task.title}',
                          form=form,
                          task=task,
                          schedule_type='mensual')

@tasks_bp.route('/tasks/<int:task_id>/biweekly-schedule', methods=['GET', 'POST'])
@login_required
@manager_required
def configure_biweekly_schedule(task_id):
    """Configurar frecuencia quincenal para una tarea"""
    task = Task.query.get_or_404(task_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != task.location.company_id):
        flash('No tienes permiso para configurar esta tarea.', 'danger')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    # Verificar que sea una tarea quincenal
    if task.frequency != TaskFrequency.QUINCENAL:
        flash('Esta tarea no es de frecuencia quincenal.', 'warning')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    form = BiweeklyScheduleForm()
    
    if form.validate_on_submit():
        log_activity(f'Frecuencia quincenal configurada para tarea: {task.title}')
        flash('Configuración guardada correctamente.', 'success')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    return render_template('tasks/schedule_form.html',
                          title=f'Configurar Tarea Quincenal: {task.title}',
                          form=form,
                          task=task,
                          schedule_type='quincenal')

@tasks_bp.route('/tasks/<int:task_id>/edit', methods=['GET', 'POST'])
@login_required
@manager_required
def edit_task(task_id):
    """Editar una tarea existente"""
    task = Task.query.get_or_404(task_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != task.location.company_id):
        flash('No tienes permiso para editar esta tarea.', 'danger')
        return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    form = TaskForm(obj=task)
    form.location_id.choices = [(task.location.id, task.location.name)]
    
    # Cargar grupos de tareas disponibles para este local
    groups = TaskGroup.query.filter_by(location_id=task.location_id).all()
    group_choices = [(0, 'Ningún grupo')] + [(g.id, g.name) for g in groups]
    form.group_id.choices = group_choices
    
    # Establecer el grupo seleccionado actualmente
    if task.group_id:
        form.group_id.data = task.group_id
    else:
        form.group_id.data = 0
    
    # Si es una tarea personalizada, marcar los días seleccionados
    if task.frequency == TaskFrequency.PERSONALIZADA:
        # Obtener los días de la semana configurados
        weekdays = TaskWeekday.query.filter_by(task_id=task.id).all()
        days_of_week = [day.day_of_week for day in weekdays]
        
        # Marcar los checkboxes correspondientes
        form.monday.data = WeekDay.LUNES in days_of_week
        form.tuesday.data = WeekDay.MARTES in days_of_week
        form.wednesday.data = WeekDay.MIERCOLES in days_of_week
        form.thursday.data = WeekDay.JUEVES in days_of_week
        form.friday.data = WeekDay.VIERNES in days_of_week
        form.saturday.data = WeekDay.SABADO in days_of_week
        form.sunday.data = WeekDay.DOMINGO in days_of_week
    
    # Guardar la frecuencia original
    original_frequency = task.frequency
    
    if form.validate_on_submit():
        # Registrar datos antes
        current_app.logger.info(f"Editando tarea ID {task.id}. Antes: título={task.title}, desc={task.description[:20]}, prioridad={task.priority}, frecuencia={task.frequency}")
        current_app.logger.info(f"Datos del formulario: título={form.title.data}, desc={form.description.data[:20]}, prioridad={form.priority.data}, frecuencia={form.frequency.data}")
        
        # Asignar valores
        task.title = form.title.data
        task.description = form.description.data
        task.priority = TaskPriority(form.priority.data)
        task.frequency = TaskFrequency(form.frequency.data)
        task.start_date = form.start_date.data
        task.end_date = form.end_date.data
        
        # Actualizar grupo de tareas
        if form.group_id.data and form.group_id.data != 0:
            task.group_id = form.group_id.data
        else:
            task.group_id = None
        
        # Registrar datos después de la actualización
        current_app.logger.info(f"Después de actualizar: título={task.title}, desc={task.description[:20]}, prioridad={task.priority}, frecuencia={task.frequency}")
        
        # Confirmar transacción
        db.session.commit()
        
        # Verificar que los cambios se guardaron
        db.session.refresh(task)
        current_app.logger.info(f"Después de commit: título={task.title}, desc={task.description[:20]}, prioridad={task.priority}, frecuencia={task.frequency}")
        
        # Si la frecuencia es personalizada, actualizar los días de la semana
        if task.frequency == TaskFrequency.PERSONALIZADA:
            # Eliminar días existentes
            TaskWeekday.query.filter_by(task_id=task.id).delete()
            
            # Crear nuevos registros de días seleccionados
            if form.monday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.LUNES))
            if form.tuesday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.MARTES))
            if form.wednesday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.MIERCOLES))
            if form.thursday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.JUEVES))
            if form.friday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.VIERNES))
            if form.saturday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.SABADO))
            if form.sunday.data:
                db.session.add(TaskWeekday(task_id=task.id, day_of_week=WeekDay.DOMINGO))
            
            db.session.commit()
        
        log_activity(f'Tarea actualizada: {task.title}')
        
        # Si cambió la frecuencia, redirigir a configurar el horario
        if original_frequency != task.frequency:
            # Eliminar horarios anteriores
            TaskSchedule.query.filter_by(task_id=task.id).delete()
            db.session.commit()
            
            flash(f'Tarea "{task.title}" actualizada. La frecuencia ha cambiado, configura el nuevo horario.', 'success')
            
            if task.frequency == TaskFrequency.DIARIA:
                return redirect(url_for('tasks.configure_daily_schedule', task_id=task.id))
            elif task.frequency == TaskFrequency.SEMANAL:
                return redirect(url_for('tasks.configure_weekly_schedule', task_id=task.id))
            elif task.frequency == TaskFrequency.MENSUAL:
                return redirect(url_for('tasks.configure_monthly_schedule', task_id=task.id))
            elif task.frequency == TaskFrequency.QUINCENAL:
                return redirect(url_for('tasks.configure_biweekly_schedule', task_id=task.id))
        else:
            flash(f'Tarea "{task.title}" actualizada correctamente.', 'success')
            return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
    
    return render_template('tasks/task_form.html',
                          title=f'Editar Tarea: {task.title}',
                          form=form,
                          task=task,
                          location=task.location,
                          location_id=task.location_id)

@tasks_bp.route('/tasks/<int:task_id>/delete', methods=['POST'])
@login_required
@manager_required
def delete_task(task_id):
    """Eliminar una tarea"""
    try:
        # Verificar si la tarea existe
        task = Task.query.get(task_id)
        if not task:
            # La tarea ya ha sido eliminada o no existe
            flash('La tarea ya ha sido eliminada.', 'info')
            # Intentar obtener location_id de los argumentos de la solicitud
            location_id = request.args.get('location_id')
            if location_id:
                return redirect(url_for('tasks.list_tasks', location_id=location_id))
            else:
                return redirect(url_for('tasks.index'))
        
        # Verificar permisos (admin o gerente de la empresa)
        if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != task.location.company_id):
            flash('No tienes permiso para eliminar esta tarea.', 'danger')
            return redirect(url_for('tasks.list_tasks', location_id=task.location_id))
        
        location_id = task.location_id
        title = task.title
        
        # Guardar info para debugging
        debug_info = {
            "task_id": task.id,
            "task_title": task.title,
            "location_id": task.location_id,
            "frequency": task.frequency,
            "month_days": []
        }
        
        # Realizar las operaciones dentro de una transacción
        try:
            # Primero obtener los días del mes para debugging
            month_days = TaskMonthDay.query.filter_by(task_id=task.id).all()
            debug_info["month_days"] = [md.day_of_month for md in month_days]
            current_app.logger.info(f"Tarea con ID {task.id} tiene los siguientes días del mes: {debug_info['month_days']}")
            
            # IMPORTANTE: Primero eliminar las instancias de tareas, que es la restricción de clave foránea que falla
            current_app.logger.info(f"Eliminando instancias de tarea para task_id={task.id}")
            from sqlalchemy import text
            task_instances_sql = text("DELETE FROM task_instances WHERE task_id = :task_id")
            result = db.session.execute(task_instances_sql, {"task_id": task.id})
            num_deleted = result.rowcount
            current_app.logger.info(f"Instancias de tarea eliminadas: {num_deleted} registros")
            
            # Eliminar primero las programaciones
            TaskSchedule.query.filter_by(task_id=task.id).delete()
            current_app.logger.info(f"Programaciones eliminadas para tarea {task.id}")
            
            # Eliminar los completados
            TaskCompletion.query.filter_by(task_id=task.id).delete()
            current_app.logger.info(f"Completados eliminados para tarea {task.id}")
            
            # Eliminar días del mes asignados (para tareas mensuales)
            # Obtener primero todos los registros para depuración
            month_days_to_delete = TaskMonthDay.query.filter_by(task_id=task.id).all()
            current_app.logger.info(f"Se encontraron {len(month_days_to_delete)} días del mes para eliminar de la tarea {task.id}")
            
            try:
                # 1. Primero intentar con DELETE directo en SQL (evitando cascadas)
                monthdays_sql = text("DELETE FROM task_monthdays WHERE task_id = :task_id")
                result = db.session.execute(monthdays_sql, {"task_id": task.id})
                num_deleted = result.rowcount
                current_app.logger.info(f"Días del mes eliminados para tarea {task.id} con SQL directo: {num_deleted} registros")
            except Exception as sql_error:
                current_app.logger.error(f"Error al eliminar con SQL directo: {sql_error}")
                # 2. Intentar con el método de eliminación masiva de SQLAlchemy
                try:
                    num_deleted = TaskMonthDay.query.filter_by(task_id=task.id).delete(synchronize_session=False)
                    current_app.logger.info(f"Días del mes eliminados para tarea {task.id} con método masivo: {num_deleted} registros")
                except Exception as month_day_error:
                    current_app.logger.error(f"Error al eliminar TaskMonthDay: {month_day_error}")
                    # 3. Intentar eliminar individualmente como último recurso
                    deleted_count = 0
                    for md in month_days_to_delete:
                        try:
                            db.session.delete(md)
                            deleted_count += 1
                        except Exception as individual_error:
                            current_app.logger.error(f"Error al eliminar día del mes {md.id}: {individual_error}")
                    current_app.logger.info(f"Días del mes eliminados individualmente: {deleted_count} de {len(month_days_to_delete)}")
            
            # Hacer commit parcial después de eliminar todas las dependencias
            db.session.commit()
            current_app.logger.info(f"Commit parcial realizado después de eliminar dependencias")
            
            # Registrar el estado antes de eliminar la tarea
            current_app.logger.info(f"Preparando eliminación final de la tarea {task.id}")
            
            # Comprobar explícitamente si aún hay monthdays asociados
            remaining_monthdays = TaskMonthDay.query.filter_by(task_id=task.id).all()
            if remaining_monthdays:
                current_app.logger.warning(f"Todavía hay {len(remaining_monthdays)} registros TaskMonthDay para la tarea {task.id}")
                # Eliminar manualmente para asegurar que se eliminan
                for rm in remaining_monthdays:
                    current_app.logger.info(f"Eliminando manualmente día del mes ID {rm.id}")
                    db.session.delete(rm)
                # Commit parcial
                db.session.commit()
            
            # Verificar si todavía hay instancias de tarea
            remaining_instances = db.session.execute(
                text("SELECT COUNT(*) as count FROM task_instances WHERE task_id = :task_id"), 
                {"task_id": task.id}
            ).scalar()
            
            if remaining_instances > 0:
                current_app.logger.warning(f"Todavía hay {remaining_instances} instancias de tarea para task_id={task.id}")
                # Eliminar de nuevo con SQL directo
                db.session.execute(task_instances_sql, {"task_id": task.id})
                db.session.commit()
            
            # Finalmente eliminar la tarea
            current_app.logger.info(f"Eliminando la tarea {task.id}")
            db.session.delete(task)
            current_app.logger.info(f"Haciendo commit final")
            db.session.commit()
            current_app.logger.info(f"Tarea {task.id} eliminada completamente")
            
            log_activity(f'Tarea eliminada: {title}')
            flash(f'Tarea "{title}" eliminada correctamente.', 'success')
            return redirect(url_for('tasks.list_tasks', location_id=location_id))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error en transacción delete_task: {e}")
            raise
            
    except Exception as main_error:
        import traceback
        error_traceback = traceback.format_exc()
        current_app.logger.error(f"Error general en delete_task: {main_error}")
        current_app.logger.error(f"Traceback completo: {error_traceback}")
        flash(f'Error al eliminar la tarea: {str(main_error)}', 'danger')
        return redirect(url_for('tasks.index'))

@tasks_bp.route('/tasks/<int:task_id>')
@login_required
def view_task(task_id):
    """Ver detalles de una tarea"""
    task = Task.query.get_or_404(task_id)
    
    # Verificar permisos (admin, gerente o empleado de la empresa)
    if not current_user.is_admin() and not (current_user.company_id == task.location.company_id):
        flash('No tienes permiso para ver esta tarea.', 'danger')
        return redirect(url_for('tasks.index'))
    
    # Obtener programación
    schedule = TaskSchedule.query.filter_by(task_id=task.id).first()
    
    # Obtener historial de completados
    completions = (db.session.query(TaskCompletion, LocalUser)
                  .join(LocalUser, TaskCompletion.local_user_id == LocalUser.id)
                  .filter(TaskCompletion.task_id == task.id)
                  .order_by(TaskCompletion.completion_date.desc())
                  .all())
    
    return render_template('tasks/task_detail.html',
                          title=f'Tarea: {task.title}',
                          task=task,
                          schedule=schedule,
                          completions=completions)

# Portal de acceso para usuarios locales
@tasks_bp.route('/portal')
def portal_selection():
    """Página de selección de portal"""
    # Añadir logging para diagnóstico
    print("[DEBUG] Accediendo a la selección de portal")
    try:
        # Obtener todos los locations disponibles
        locations = Location.query.filter_by(is_active=True).all()
        print(f"[DEBUG] Locales encontrados: {len(locations)}")
        return render_template('tasks/portal_selection.html',
                            title='Selección de Portal',
                            locations=locations)
    except Exception as e:
        print(f"[ERROR] Error en portal_selection: {str(e)}")
        # Devolver una respuesta mínima para evitar 500
        return f"Error: {str(e)}", 500

@tasks_bp.route('/portal-test')
def portal_test():
    """Ruta de prueba para diagnóstico"""
    try:
        # Intentamos primero una consulta simple sin usar los campos nuevos
        locations = db.session.query(Location.id, Location.name).filter_by(is_active=True).all()
        location_count = len(locations)
        location_names = [loc[1] for loc in locations]
        
        # Ahora verificamos los campos nuevos
        has_new_columns = True
        try:
            test_query = db.session.query(Location.id, Location.portal_username).first()
        except Exception as e:
            has_new_columns = False
            column_error = str(e)
        
        # Generar respuesta informativa
        result = {
            "status": "ok",
            "locations_count": location_count,
            "locations": location_names,
            "has_new_columns": has_new_columns
        }
        
        if not has_new_columns:
            result["column_error"] = column_error
            
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@tasks_bp.route('/portal/<int:location_id>', methods=['GET', 'POST'])
def portal_login(location_id):
    """Página de login para acceder al portal de un local"""
    location = Location.query.get_or_404(location_id)
    
    form = PortalLoginForm()
    
    # Mostrar información de las credenciales fijas al cargar la página
    if not request.method == 'POST':
        flash(f"Credenciales fijas para este portal: Usuario: {location.portal_fixed_username}, Contraseña: {location.portal_fixed_password}", 
              "info")
    
    if form.validate_on_submit():
        # Verificar el nombre de usuario
        if form.username.data == location.portal_fixed_username:
            # Verificar la contraseña usando el método de verificación específico
            if location.check_portal_password(form.password.data):
                # Guardar en sesión que estamos autenticados en este portal
                session['portal_authenticated'] = True
                session['portal_location_id'] = location.id
                
                # Redireccionar al portal real
                return redirect(url_for('tasks.local_portal', location_id=location.id))
            else:
                flash('Contraseña incorrecta.', 'danger')
        else:
            flash('Nombre de usuario incorrecto.', 'danger')
    
    return render_template('tasks/portal_login.html',
                          title=f'Acceso al Portal {location.name}',
                          location=location,
                          form=form)

@tasks_bp.route('/local-login', methods=['GET', 'POST'])
def local_login():
    """Página de login para usuarios locales"""
    # Verificar si ya hay un usuario local en la sesión
    if 'local_user_id' in session:
        return redirect(url_for('tasks.local_user_labels'))
    
    # Obtener todos los locales activos
    locations = Location.query.filter_by(is_active=True).all()
    
    # Si hay un local guardado en la sesión, usarlo
    saved_location_id = session.get('saved_location_id')
    
    # Si solo hay un local, redireccionar al portal
    if len(locations) == 1:
        location = locations[0]
        return redirect(url_for('tasks.local_portal', location_id=location.id))
    
    return render_template('tasks/local_login.html',
                          title='Acceso a Portal de Tareas',
                          locations=locations,
                          saved_location_id=saved_location_id)

@tasks_bp.route('/local-portal/<int:location_id>')
def local_portal(location_id):
    """Portal de acceso para un local"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar si el usuario está autenticado en el portal
    if not session.get('portal_authenticated') or session.get('portal_location_id') != location.id:
        # Si no está autenticado, redireccionar al login del portal
        return redirect(url_for('tasks.portal_login', location_id=location.id))
    
    # Guardar ID del local en la sesión
    session['location_id'] = location_id
    
    # Obtener usuarios activos del local
    local_users = LocalUser.query.filter_by(location_id=location_id, is_active=True).all()
    
    # Si no hay usuarios, crear un admin por defecto
    if not local_users:
        try:
            # Crear usuario admin por defecto
            admin_user = LocalUser(
                name="Admin",
                last_name="Local",
                username=f"admin_{location_id}",
                is_active=True,
                location_id=location_id
            )
            
            # Establecer PIN 1234
            admin_user.set_pin("1234")
            
            db.session.add(admin_user)
            db.session.commit()
            
            # Refrescar la lista de usuarios
            local_users = [admin_user]
            
            # Mostrar mensaje informativo 
            flash("Se ha creado un usuario administrador por defecto con PIN: 1234", "info")
            
        except Exception as e:
            flash(f"No se pudo crear el usuario por defecto: {str(e)}", "warning")
    
    return render_template('tasks/local_portal.html',
                          title=f'Portal de {location.name}',
                          location=location,
                          local_users=local_users)

@tasks_bp.route('/local-user-login/<int:user_id>', methods=['GET', 'POST'])
def local_user_login(user_id):
    """Login con PIN para empleado local"""
    if 'location_id' not in session:
        return redirect(url_for('tasks.index'))
    
    user = LocalUser.query.get_or_404(user_id)
    
    # Verificar que pertenece al local correcto
    if user.location_id != session['location_id']:
        flash('Usuario no válido para este local.', 'danger')
        return redirect(url_for('tasks.local_portal', location_id=session['location_id']))
    
    # Obtener configuración de la ubicación
    location = Location.query.get(session['location_id'])
    
    # Si la ubicación no requiere PIN, permitir acceso directo
    if location and not location.requires_pin:
        # Guardar usuario en sesión sin verificar PIN
        session['local_user_id'] = user.id
        session['local_user_name'] = f"{user.name} {user.last_name}"
        
        log_activity(f'Acceso de usuario local (sin PIN): {user.name} en {user.location.name}')
        flash(f'Bienvenido, {user.name}!', 'success')
        return redirect(url_for('tasks.local_user_tasks'))
    
    # Si requiere PIN, continuar con la verificación normal
    form = LocalUserPinForm()
    
    if form.validate_on_submit():
        if user.check_pin(form.pin.data):
            # Guardar usuario en sesión
            session['local_user_id'] = user.id
            session['local_user_name'] = f"{user.name} {user.last_name}"
            
            log_activity(f'Acceso de usuario local: {user.name} en {user.location.name}')
            flash(f'Bienvenido, {user.name}!', 'success')
            return redirect(url_for('tasks.local_user_tasks'))
        else:
            flash('PIN incorrecto.', 'danger')
    
    return render_template('tasks/local_user_pin.html',
                          title=f'Acceso de {user.name}',
                          form=form,
                          user=user)

@tasks_bp.route('/local-logout')
def local_logout():
    """Cerrar sesión de usuario local"""
    location_id = session.get('location_id')
    session.pop('local_user_id', None)
    session.pop('local_user_name', None)
    flash('Has cerrado sesión correctamente.', 'success')
    
    if location_id:
        return redirect(url_for('tasks.local_portal', location_id=location_id))
    else:
        return redirect(url_for('tasks.index'))

@tasks_bp.route('/portal-logout')
def portal_logout():
    """Cerrar sesión de portal local"""
    # Limpiar todas las variables de sesión relacionadas con el portal
    session.pop('location_id', None)
    session.pop('local_user_username', None)
    session.pop('local_user_id', None)
    session.pop('local_user_name', None)
    session.pop('portal_authenticated', None)
    session.pop('portal_location_id', None)
    
    flash('Has cerrado sesión del portal correctamente.', 'success')
    return redirect(url_for('tasks.index'))

import fcntl
import os

# Archivos de bloqueo para evitar ejecuciones simultáneas de pruebas manuales
TEST_WEEKLY_RESET_LOCK_FILE = "/tmp/test_weekly_reset_tasks.lock"
TEST_DAILY_RESET_LOCK_FILE = "/tmp/test_daily_reset_tasks.lock"

@tasks_bp.route('/local-user/test-reset-weekly-tasks', methods=['POST'])
@local_user_required
def test_reset_weekly_tasks():
    """Ejecuta manualmente el reinicio de tareas semanales (solo para pruebas)"""
    # Usar un archivo de bloqueo para evitar ejecuciones simultáneas
    lock_file = None
    try:
        # Intentar obtener un bloqueo exclusivo (no bloqueante)
        lock_file = open(TEST_WEEKLY_RESET_LOCK_FILE, 'w')
        try:
            fcntl.lockf(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
            
            # Escribir información en el archivo de bloqueo
            pid = os.getpid()
            lock_info = f"{pid} {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            lock_file.write(lock_info)
            lock_file.flush()
            
            # Si llegamos aquí, tenemos el bloqueo exclusivo
            # Ejecutar manualmente el reinicio de tareas semanales
            tasks_reset_count = reset_weekly_tasks()
            
            # Procesar tareas personalizadas para la semana
            custom_tasks_count = process_custom_tasks_for_week()
            
            log_activity(f'Reinicio manual de tareas semanales ejecutado: {tasks_reset_count} tareas reiniciadas, {custom_tasks_count} instancias creadas')
            flash('¡Reinicio de tareas semanales ejecutado correctamente!', 'success')
            
        except IOError:
            # No se pudo obtener el bloqueo, otro proceso ya lo tiene
            flash('Ya se está ejecutando un reinicio de tareas. Por favor, espera unos segundos antes de intentarlo de nuevo.', 'warning')
            
    except Exception as e:
        flash(f'Error al ejecutar el reinicio de tareas semanales: {str(e)}', 'danger')
    finally:
        # Liberar el bloqueo si lo conseguimos
        if lock_file:
            try:
                fcntl.lockf(lock_file, fcntl.LOCK_UN)
                lock_file.close()
            except:
                pass
    
    # Redirigir a la página de tareas
    return redirect(url_for('tasks.local_user_tasks'))

@tasks_bp.route('/local-user/test-reset-daily-tasks', methods=['POST'])
@local_user_required
def test_reset_daily_tasks():
    """Ejecuta manualmente el reinicio de tareas diarias y mensuales (solo para pruebas)"""
    # Importar las funciones de reinicio de tareas
    from daily_tasks_reset_service import reset_daily_tasks, reset_monthly_tasks
    
    # Usar un archivo de bloqueo para evitar ejecuciones simultáneas
    lock_file = None
    try:
        # Intentar obtener un bloqueo exclusivo (no bloqueante)
        lock_file = open(TEST_DAILY_RESET_LOCK_FILE, 'w')
        try:
            fcntl.lockf(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
            
            # Escribir información en el archivo de bloqueo
            pid = os.getpid()
            lock_info = f"{pid} {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            lock_file.write(lock_info)
            lock_file.flush()
            
            # Si llegamos aquí, tenemos el bloqueo exclusivo
            # Ejecutar manualmente el reinicio de tareas diarias
            daily_tasks_count = reset_daily_tasks()
            
            # Ejecutar manualmente el reinicio de tareas mensuales
            monthly_tasks_count = reset_monthly_tasks()
            
            # Registrar la actividad en el log y mostrar mensaje
            log_activity(f'Reinicio manual ejecutado: {daily_tasks_count} tareas diarias y {monthly_tasks_count} tareas mensuales actualizadas')
            flash(f'¡Reinicio de tareas ejecutado correctamente! ({daily_tasks_count} tareas diarias, {monthly_tasks_count} tareas mensuales)', 'success')
            
        except IOError:
            # No se pudo obtener el bloqueo, otro proceso ya lo tiene
            flash('Ya se está ejecutando un reinicio de tareas. Por favor, espera unos segundos antes de intentarlo de nuevo.', 'warning')
            
    except Exception as e:
        flash(f'Error al ejecutar el reinicio de tareas diarias: {str(e)}', 'danger')
    finally:
        # Liberar el bloqueo si lo conseguimos
        if lock_file:
            try:
                fcntl.lockf(lock_file, fcntl.LOCK_UN)
                lock_file.close()
            except:
                pass
    
    # Redirigir a la página de tareas
    return redirect(url_for('tasks.local_user_tasks'))

# Para mantener compatibilidad, redirigir la ruta antigua a la nueva
@tasks_bp.route('/local-user/test-reset-tasks', methods=['POST'])
@local_user_required
def test_reset_tasks():
    """Ruta anterior para compatibilidad"""
    return redirect(url_for('tasks.test_reset_weekly_tasks'))

@tasks_bp.route('/purge-monthly-tasks/<int:location_id>')
@login_required
def purge_monthly_tasks(location_id):
    """Elimina todas las tareas mensuales y sus registros asociados para una ubicación"""
    # Verificar permisos (sólo admin)
    if not current_user.is_admin():
        flash('Solo administradores pueden usar esta función de mantenimiento.', 'danger')
        return redirect(url_for('tasks.list_tasks', location_id=location_id))
    
    results = {
        'tasks_found': 0,
        'tasks_deleted': 0,
        'completions_deleted': 0,
        'schedules_deleted': 0,
        'month_days_deleted': 0,
        'errors': []
    }
    
    try:
        # Obtener todas las tareas mensuales de esta ubicación
        monthly_tasks = Task.query.filter_by(
            location_id=location_id, 
            frequency=TaskFrequency.MENSUAL
        ).all()
        
        task_count = len(monthly_tasks)
        results['tasks_found'] = task_count
        
        if task_count == 0:
            flash('No se encontraron tareas mensuales para purgar.', 'info')
            return redirect(url_for('tasks.list_tasks', location_id=location_id))
        
        # Registro para análisis
        current_app.logger.info(f"Iniciando purga de {task_count} tareas mensuales en ubicación {location_id}")
        
        # ESTRATEGIA: Recolectar todas las IDs primero
        task_ids = [task.id for task in monthly_tasks]
        
        try:
            # PASO 1: Eliminar completados (operación relativamente segura)
            from sqlalchemy import text
            # Usar SQL directo para evitar problemas de cascada
            completion_sql = text("DELETE FROM task_completions WHERE task_id IN :task_ids")
            result = db.session.execute(completion_sql, {"task_ids": tuple(task_ids) if len(task_ids) > 1 else f"({task_ids[0]})"})
            completions_deleted = result.rowcount
            current_app.logger.info(f"Eliminados {completions_deleted} registros de completado de tareas mensuales")
            results['completions_deleted'] = completions_deleted
            
            # PASO 2: Eliminar programación (también debería ser seguro)
            schedule_sql = text("DELETE FROM task_schedules WHERE task_id IN :task_ids")
            result = db.session.execute(schedule_sql, {"task_ids": tuple(task_ids) if len(task_ids) > 1 else f"({task_ids[0]})"})
            schedules_deleted = result.rowcount
            current_app.logger.info(f"Eliminados {schedules_deleted} registros de programación de tareas mensuales")
            results['schedules_deleted'] = schedules_deleted
            
            # PASO 3: Eliminar días del mes (SQL directo para evitar problemas de relación)
            monthday_sql = text("DELETE FROM task_monthdays WHERE task_id IN :task_ids")
            result = db.session.execute(monthday_sql, {"task_ids": tuple(task_ids) if len(task_ids) > 1 else f"({task_ids[0]})"})
            monthdays_deleted = result.rowcount
            current_app.logger.info(f"Eliminados {monthdays_deleted} registros de días mensuales")
            results['month_days_deleted'] = monthdays_deleted
            
            # Si todo ha ido bien hasta aquí, intentamos eliminar las tareas principales
            db.session.commit()
            current_app.logger.info("Commit de eliminación de registros relacionados realizado correctamente")
            
        except Exception as batch_error:
            db.session.rollback()
            current_app.logger.error(f"Error en eliminación por lotes: {batch_error}. Cambiando a modo individual...")
            results['errors'].append(f"Error en eliminación por lotes: {str(batch_error)}")
        
        # PASO 4: Eliminar cada tarea mensual individualmente 
        deleted_count = 0
        for task in monthly_tasks:
            try:
                # Registramos datos de la tarea antes de eliminarla
                task_id = task.id
                task_title = task.title
                
                # Verificar si aún hay registros relacionados (por si la eliminación en lote falló)
                if TaskCompletion.query.filter_by(task_id=task_id).count() > 0:
                    TaskCompletion.query.filter_by(task_id=task_id).delete()
                    
                if TaskSchedule.query.filter_by(task_id=task_id).count() > 0:
                    TaskSchedule.query.filter_by(task_id=task_id).delete()
                    
                if TaskMonthDay.query.filter_by(task_id=task_id).count() > 0:
                    TaskMonthDay.query.filter_by(task_id=task_id).delete()
                
                # Finalmente, eliminar la tarea
                db.session.delete(task)
                
                # Intentar commit por cada tarea para garantizar que al menos algunas se eliminen
                db.session.commit()
                
                # Registrar éxito para esta tarea
                current_app.logger.info(f"Tarea mensual {task_id} - {task_title} eliminada correctamente")
                deleted_count += 1
                
            except Exception as task_error:
                db.session.rollback()
                error_msg = f"Error al eliminar tarea mensual {task.id}: {task_error}"
                current_app.logger.error(error_msg)
                results['errors'].append(error_msg)
                # Continuar con la siguiente tarea
        
        results['tasks_deleted'] = deleted_count
        
        # Mensaje de resultado
        if deleted_count == task_count:
            flash(f'Se eliminaron correctamente todas las {deleted_count} tareas mensuales.', 'success')
        else:
            flash(f'Se eliminaron {deleted_count} de {task_count} tareas mensuales. Revise los logs para más detalles.', 'warning')
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_traceback = traceback.format_exc()
        current_app.logger.error(f"Error general en purge_monthly_tasks: {e}")
        current_app.logger.error(f"Traceback completo: {error_traceback}")
        flash(f'Error al purgar tareas mensuales: {str(e)}', 'danger')
        results['errors'].append(str(e))
    
    # Registrar actividad
    log_activity(f'Purga de tareas mensuales ejecutada. Resultados: {results}')
    return redirect(url_for('tasks.list_tasks', location_id=location_id))

@tasks_bp.route('/fix-task-monthdays/<int:location_id>')
@login_required
def fix_task_monthdays(location_id):
    """Repara o elimina registros huérfanos de TaskMonthDay y corrige referencias incorrectas"""
    # Verificar permisos (sólo admin)
    if not current_user.is_admin():
        flash('Solo administradores pueden usar esta función de mantenimiento.', 'danger')
        return redirect(url_for('tasks.list_tasks', location_id=location_id))
    
    results = {
        'orphaned_deleted': 0,
        'invalid_days_fixed': 0,
        'duplicate_days_fixed': 0,
        'errors': []
    }
    
    try:
        # Obtener todas las tareas de esta ubicación y un diccionario de ids para verificación rápida
        tasks = Task.query.filter_by(location_id=location_id).all()
        task_ids = {task.id: task for task in tasks}
        
        current_app.logger.info(f"Iniciando corrección de TaskMonthDay - Encontradas {len(tasks)} tareas en ubicación {location_id}")
        
        # PASO 1: Encontrar y eliminar registros TaskMonthDay huérfanos (sin tarea asociada)
        orphaned_records = TaskMonthDay.query.filter(
            ~TaskMonthDay.task_id.in_(task_ids.keys()) if task_ids else True
        ).all()
        
        orphan_count = len(orphaned_records)
        current_app.logger.info(f"Se encontraron {orphan_count} registros TaskMonthDay huérfanos")
        
        if orphan_count > 0:
            # Mostrar detalles de registros huérfanos antes de eliminar
            for record in orphaned_records:
                current_app.logger.info(f"Eliminando registro huérfano: ID={record.id}, task_id={record.task_id}, día={record.day_of_month}")
                db.session.delete(record)
            
            results['orphaned_deleted'] = orphan_count
        
        # PASO 2: Verificar y corregir días del mes inválidos
        all_month_days = TaskMonthDay.query.filter(
            TaskMonthDay.task_id.in_(task_ids.keys())
        ).all()
        
        invalid_days = [md for md in all_month_days if md.day_of_month < 1 or md.day_of_month > 31]
        if invalid_days:
            current_app.logger.warning(f"Se encontraron {len(invalid_days)} registros con días del mes inválidos")
            for invalid in invalid_days:
                current_app.logger.warning(f"Corrigiendo día inválido: ID={invalid.id}, task_id={invalid.task_id}, día={invalid.day_of_month}")
                # Corregir valores fuera de rango
                if invalid.day_of_month < 1:
                    invalid.day_of_month = 1
                elif invalid.day_of_month > 31:
                    invalid.day_of_month = 31
            
            results['invalid_days_fixed'] = len(invalid_days)
        
        # PASO 3: Eliminar duplicados (misma tarea, mismo día)
        from sqlalchemy import func
        duplicates = db.session.query(
            TaskMonthDay.task_id,
            TaskMonthDay.day_of_month,
            func.count(TaskMonthDay.id).label('count')
        ).group_by(
            TaskMonthDay.task_id,
            TaskMonthDay.day_of_month
        ).having(
            func.count(TaskMonthDay.id) > 1
        ).all()
        
        duplicate_count = 0
        for dup in duplicates:
            task_id, day_of_month, count = dup
            current_app.logger.warning(f"Encontrados {count} registros duplicados para tarea {task_id}, día {day_of_month}")
            
            # Obtener todos los registros duplicados excepto el primero (que mantendremos)
            dup_records = TaskMonthDay.query.filter_by(
                task_id=task_id,
                day_of_month=day_of_month
            ).order_by(TaskMonthDay.id).offset(1).all()
            
            for dr in dup_records:
                current_app.logger.info(f"Eliminando duplicado: ID={dr.id}, task_id={dr.task_id}, día={dr.day_of_month}")
                db.session.delete(dr)
                duplicate_count += 1
        
        results['duplicate_days_fixed'] = duplicate_count
        
        # Confirmar cambios solo si hay algo que guardar
        if results['orphaned_deleted'] > 0 or results['invalid_days_fixed'] > 0 or results['duplicate_days_fixed'] > 0:
            db.session.commit()
            
            # Mensaje con detalles de las correcciones
            message_parts = []
            if results['orphaned_deleted'] > 0:
                message_parts.append(f"{results['orphaned_deleted']} registros huérfanos eliminados")
            if results['invalid_days_fixed'] > 0:
                message_parts.append(f"{results['invalid_days_fixed']} días inválidos corregidos")
            if results['duplicate_days_fixed'] > 0:
                message_parts.append(f"{results['duplicate_days_fixed']} registros duplicados eliminados")
                
            flash(f"Corrección completada: {', '.join(message_parts)}.", 'success')
        else:
            flash('No se encontraron problemas que corregir en los registros TaskMonthDay.', 'info')
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_traceback = traceback.format_exc()
        current_app.logger.error(f"Error en fix_task_monthdays: {e}")
        current_app.logger.error(f"Traceback completo: {error_traceback}")
        flash(f'Error al reparar registros: {str(e)}', 'danger')
        results['errors'].append(str(e))
    
    # Registrar actividad
    log_activity(f'Herramienta de corrección TaskMonthDay ejecutada. Resultados: {results}')
    return redirect(url_for('tasks.list_tasks', location_id=location_id))


@tasks_bp.route('/local-user/tasks', defaults={'date_str': None, 'group_id': None})
@tasks_bp.route('/local-user/tasks/<date_str>', defaults={'group_id': None})
@tasks_bp.route('/local-user/tasks/group/<int:group_id>', defaults={'date_str': None})
@tasks_bp.route('/local-user/tasks/<date_str>/group/<int:group_id>')
@local_user_required
def local_user_tasks(date_str=None, group_id=None):
    """Panel de tareas para usuario local"""
    user_id = session['local_user_id']
    user = LocalUser.query.get_or_404(user_id)
    location = user.location
    
    # Determinar la fecha de las tareas (hoy por defecto, o la especificada)
    today = date.today()
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = today
    else:
        selected_date = today
    
    # Calcular fechas para el carrusel (día anterior, actual, siguiente)
    prev_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1)
    
    # Determinar los nombres de los días en español (abreviado para mostrar)
    days_map_short = {
        0: 'LUN',
        1: 'MAR',
        2: 'MIÉ',
        3: 'JUE',
        4: 'VIE', 
        5: 'SÁB',
        6: 'DOM'
    }
    
    # Mapeo de día de la semana a valor de enum WeekDay
    days_map_enum = {
        0: WeekDay.LUNES,
        1: WeekDay.MARTES,
        2: WeekDay.MIERCOLES,
        3: WeekDay.JUEVES,
        4: WeekDay.VIERNES,
        5: WeekDay.SABADO,
        6: WeekDay.DOMINGO
    }
    
    # Configurar el carrusel de fechas, preservando el filtro de grupo si existe
    base_url_params = {}
    if group_id is not None:
        base_url_params['group_id'] = group_id
        
    date_carousel = [
        {
            'date': prev_date,
            'day_name': days_map_short[prev_date.weekday()],
            'day': prev_date.day,
            'url': url_for('tasks.local_user_tasks', date_str=prev_date.strftime('%Y-%m-%d'), **base_url_params),
            'active': False
        },
        {
            'date': selected_date,
            'day_name': days_map_short[selected_date.weekday()],
            'day': selected_date.day,
            'url': url_for('tasks.local_user_tasks', date_str=selected_date.strftime('%Y-%m-%d'), **base_url_params),
            'active': True
        },
        {
            'date': next_date, 
            'day_name': days_map_short[next_date.weekday()],
            'day': next_date.day,
            'url': url_for('tasks.local_user_tasks', date_str=next_date.strftime('%Y-%m-%d'), **base_url_params),
            'active': False
        }
    ]
    
    # Inicializar colecciones
    active_tasks = []
    grouped_tasks = {}  # Diccionario para agrupar tareas por grupo
    ungrouped_tasks = [] # Lista para tareas sin grupo
    
    # Aplicar filtro de grupo si es necesario
    tasks_query = Task.query.filter_by(location_id=location.id, status=TaskStatus.PENDIENTE)
    
    # Si hay un grupo_id especificado, filtrar por ese grupo
    if group_id is not None:
        if group_id == 0:  # Tareas sin grupo
            tasks_query = tasks_query.filter(Task.group_id == None)
        else:  # Tareas del grupo específico
            tasks_query = tasks_query.filter(Task.group_id == group_id)
    
    # Obtener todas las tareas pendientes (aplicando los filtros)
    pending_tasks = tasks_query.all()
    
    # Obtener grupos de tareas para esta ubicación
    task_groups = TaskGroup.query.filter_by(location_id=location.id).all()
    group_dict = {group.id: group for group in task_groups}
    
    # Obtener el grupo actual si se especificó
    current_group = None
    if group_id and group_id > 0:
        current_group = TaskGroup.query.get(group_id)
    
    # Obtener todas las completadas de la fecha seleccionada
    all_completions = db.session.query(
        TaskCompletion, LocalUser
    ).join(
        LocalUser, TaskCompletion.local_user_id == LocalUser.id
    ).filter(
        TaskCompletion.task_id.in_([t.id for t in pending_tasks]),
        db.func.date(TaskCompletion.completion_date) == selected_date
    ).order_by(
        TaskCompletion.completion_date.desc()
    ).all()
    
    # Crear un diccionario de información de completado por tarea
    completion_info = {}
    for completion, local_user in all_completions:
        completion_info[completion.task_id] = {
            'user': f"{local_user.name} {local_user.last_name}",
            'time': completion.completion_date.strftime('%H:%M'),
            'completed_by_me': local_user.id == user_id
        }
    
    # Completiones específicas para este usuario en la fecha seleccionada
    user_completions = TaskCompletion.query.filter_by(
        local_user_id=user_id
    ).filter(
        db.func.date(TaskCompletion.completion_date) == selected_date
    ).order_by(
        TaskCompletion.completion_date.desc()
    ).all()
    
    # Lista de IDs de tareas completadas en la fecha seleccionada
    completed_task_ids = [c.task_id for c, _ in all_completions]
    
    # Función auxiliar para comprobar si una tarea debe aparecer en la fecha seleccionada
    def task_is_due_on_date(task, check_date):
        # Si la fecha está fuera del rango de fechas de la tarea, no es debido
        if task.start_date and check_date < task.start_date:
            return False
        if task.end_date and check_date > task.end_date:
            return False
        
        # Verificar frecuencia
        if task.frequency == TaskFrequency.DIARIA:
            return True
        
        # Para tareas semanales, verificar día de la semana y estado de completado semanal
        if task.frequency == TaskFrequency.SEMANAL:
            # Si está completada para esta semana, no mostrarla
            if task.current_week_completed and check_date.weekday() < 7:  # cualquier día de la semana
                return False
                
            weekday_name = days_map_enum[check_date.weekday()]
            for schedule in task.schedule_details:
                if schedule.day_of_week and schedule.day_of_week.value == weekday_name.value:
                    return True
            return False
        
        # Para tareas quincenales, verificar quincena
        if task.frequency == TaskFrequency.QUINCENAL:
            start_date = task.start_date or task.created_at.date()
            days_diff = (check_date - start_date).days
            return days_diff % 14 == 0
        
        # Para tareas mensuales, verificar día del mes
        if task.frequency == TaskFrequency.MENSUAL:
            # Si la tarea ya está completada para el mes actual, no mostrarla
            if task.current_month_completed:
                return False
                
            # Si hay días específicos configurados con el nuevo modelo
            if task.monthdays:
                # Verificar si el día actual coincide con alguno de los días configurados
                return any(monthday.day_of_month == check_date.day for monthday in task.monthdays)
                
            # Si no hay días específicos configurados con el nuevo modelo,
            # intentar con el modelo antiguo (TaskSchedule)
            schedules = [s for s in task.schedule_details if s.day_of_month]
            if schedules:
                return any(s.day_of_month == check_date.day for s in schedules)
                
            # Si no hay horario específico, usar el día de inicio como referencia
            start_date = task.start_date or task.created_at.date()
            return start_date.day == check_date.day
        
        # Para tareas personalizadas, verificar días específicos
        if task.frequency == TaskFrequency.PERSONALIZADA:
            # Verificar si alguno de los días de la semana coincide
            weekday_value = days_map_enum[check_date.weekday()].value
            for weekday in task.weekdays:
                if weekday.day_of_week.value == weekday_value:
                    return True
            return False
        
        return False
    
    for task in pending_tasks:
        # Usar la nueva función para verificar tareas en fechas específicas
        task_due = False
        if selected_date == today:
            # Para hoy, usar el método existente por compatibilidad
            task_due = task.is_due_today()
        else:
            # Para otras fechas, usar nuestra nueva lógica
            task_due = task_is_due_on_date(task, selected_date)
            
        if task_due:
            # Verificar si ya ha sido completada
            task.completed_today = task.id in completed_task_ids
            
            # Agregar información de quién completó la tarea
            if task.id in completion_info:
                task.completion_info = completion_info[task.id]
            
            active_tasks.append(task)
            
            # Agrupar por grupo si tiene uno
            if task.group_id and task.group_id in group_dict:
                task_group_id = task.group_id
                if task_group_id not in grouped_tasks:
                    grouped_tasks[task_group_id] = {
                        'group': group_dict[task_group_id],
                        'tasks': []
                    }
                grouped_tasks[task_group_id]['tasks'].append(task)
            else:
                ungrouped_tasks.append(task)
    
    # Personalizar título en base al filtro
    if current_group:
        title = f'Tareas de {current_group.name} - {user.name}'
    elif group_id == 0:
        title = f'Tareas sin clasificar - {user.name}'
    else:
        title = f'Tareas de {user.name}'
        
    return render_template('tasks/local_user_tasks.html',
                          title=title,
                          user=user,
                          local_user=user,
                          location=location,
                          active_tasks=active_tasks,
                          ungrouped_tasks=ungrouped_tasks,
                          grouped_tasks=grouped_tasks,
                          task_groups=task_groups,
                          completed_tasks=user_completions,
                          selected_date=selected_date,
                          today=today, 
                          date_carousel=date_carousel,
                          current_group=current_group,
                          group_id=group_id)

@tasks_bp.route('/local-user/tasks/<int:task_id>/complete', methods=['GET', 'POST'])
@local_user_required
def complete_task(task_id):
    """Marcar una tarea como completada (versión con formulario)"""
    user_id = session['local_user_id']
    user = LocalUser.query.get_or_404(user_id)
    task = Task.query.get_or_404(task_id)
    
    # Verificar que la tarea pertenece al local del usuario
    if task.location_id != user.location_id:
        flash('Tarea no válida para este local.', 'danger')
        return redirect(url_for('tasks.local_user_tasks'))
    
    # Verificar si ya ha sido completada hoy por este usuario
    today = date.today()
    completion = TaskCompletion.query.filter_by(
        task_id=task.id,
        local_user_id=user_id
    ).filter(
        db.func.date(TaskCompletion.completion_date) == today
    ).first()
    
    if completion:
        flash('Ya has completado esta tarea hoy.', 'warning')
        return redirect(url_for('tasks.local_user_tasks'))
    
    form = TaskCompletionForm()
    
    if form.validate_on_submit():
        completion = TaskCompletion(
            task_id=task.id,
            local_user_id=user_id,
            notes=form.notes.data
        )
        
        # Actualizar el estado de la tarea a completada
        task.status = TaskStatus.COMPLETADA
        
        # Si es una tarea semanal, marcar como completada para esta semana
        if task.frequency == TaskFrequency.SEMANAL:
            task.current_week_completed = True
            
        # Si es una tarea mensual, marcar como completada para este mes
        elif task.frequency == TaskFrequency.MENSUAL:
            task.current_month_completed = True
        
        db.session.add(completion)
        db.session.commit()
        
        log_activity(f'Tarea completada: {task.title} por {user.name}')
        flash('¡Tarea completada correctamente!', 'success')
        return redirect(url_for('tasks.local_user_tasks'))
    
    return render_template('tasks/complete_task.html',
                          title=f'Completar Tarea: {task.title}',
                          form=form,
                          task=task,
                          user=user)

@tasks_bp.route('/local-user/tasks/<int:task_id>/ajax-complete', methods=['POST'])
@local_user_required
def ajax_complete_task(task_id):
    """Marcar una tarea como completada (versión AJAX)"""
    user_id = session['local_user_id']
    user = LocalUser.query.get_or_404(user_id)
    task = Task.query.get_or_404(task_id)
    
    # Verificar que la solicitud es AJAX
    if not request.is_json:
        return jsonify({'error': 'Se requiere petición JSON'}), 400
    
    # Verificar que la tarea pertenece al local del usuario
    if task.location_id != user.location_id:
        return jsonify({'error': 'Tarea no válida para este local'}), 403
    
    # Verificar si ya ha sido completada hoy por este usuario
    today = date.today()
    existing_completion = TaskCompletion.query.filter_by(
        task_id=task.id,
        local_user_id=user_id
    ).filter(
        db.func.date(TaskCompletion.completion_date) == today
    ).first()
    
    if existing_completion:
        return jsonify({'error': 'Ya has completado esta tarea hoy'}), 400
    
    # Obtener notas (opcional)
    data = request.json
    notes = data.get('notes', '')
    
    # Crear registro de tarea completada
    completion = TaskCompletion(
        task_id=task.id,
        local_user_id=user_id,
        notes=notes
    )
    
    # Actualizar el estado de la tarea a completada
    task.status = TaskStatus.COMPLETADA
    
    # Si es una tarea semanal, marcar como completada para esta semana
    if task.frequency == TaskFrequency.SEMANAL:
        task.current_week_completed = True
        
    # Si es una tarea mensual, marcar como completada para este mes
    elif task.frequency == TaskFrequency.MENSUAL:
        task.current_month_completed = True
    
    db.session.add(completion)
    db.session.commit()
    
    log_activity(f'Tarea completada (AJAX): {task.title} por {user.name}')
    
    # Devolver información para actualizar la UI
    return jsonify({
        'success': True,
        'taskId': task.id,
        'taskTitle': task.title,
        'completedBy': f"{user.name} {user.last_name}",
        'completedAt': datetime.now().strftime('%H:%M')
    })

# API para regenerar contraseña del portal
@tasks_bp.route('/api/regenerate-password/<int:location_id>', methods=['GET', 'POST'])
@login_required
@manager_required
def regenerate_password(location_id):
    """Devuelve la contraseña fija del portal de un local"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        return jsonify({'error': 'No tienes permiso para obtener la contraseña de este local'}), 403
    
    # Devolver la contraseña fija
    fixed_password = location.portal_fixed_password
    
    return jsonify({'success': True, 'password': fixed_password})

# Actualizar credenciales personalizadas del portal
@tasks_bp.route('/locations/<int:location_id>/update-credentials', methods=['POST'])
@login_required
@manager_required
def update_portal_credentials(location_id):
    """Actualiza las credenciales del portal para un local (nombre y contraseña)"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        flash('No tienes permiso para modificar las credenciales de este local', 'danger')
        return redirect(url_for('tasks.list_locations'))
    
    # Obtener datos del formulario
    portal_name = request.form.get('portal_name', '').strip()
    custom_username = request.form.get('custom_username', '').strip()
    custom_password = request.form.get('custom_password', '').strip()
    
    changes_made = False
    
    try:
        # Actualizar nombre del portal si se ha proporcionado y es diferente al actual
        if portal_name and portal_name != location.name:
            old_name = location.name
            location.name = portal_name
            changes_made = True
            
            # Registrar cambio en los logs
            log_activity(f'Actualización de nombre del portal para local: {old_name} → {portal_name}', user_id=current_user.id)
        
        # Actualizar nombre de usuario personalizado si se ha proporcionado uno nuevo
        if custom_username and custom_username != location.portal_fixed_username:
            old_username = location.portal_fixed_username
            try:
                location.set_portal_username(custom_username)
                changes_made = True
                
                # Registrar cambio en los logs
                log_activity(f'Actualización de nombre de usuario del portal para local: {old_username} → {custom_username}', user_id=current_user.id)
            except ValueError as e:
                flash(str(e), 'danger')
                return redirect(url_for('tasks.location_detail', id=location.id))
        
        # Actualizar contraseña personalizada solo si se ha proporcionado una nueva
        if custom_password:
            location.set_portal_password(custom_password)
            changes_made = True
            
            # Registrar cambio en los logs
            log_activity(f'Actualización de contraseña del portal para local: {location.name}', user_id=current_user.id)
        
        if changes_made:
            db.session.commit()
            flash('Credenciales del portal actualizadas correctamente', 'success')
        else:
            flash('No se han detectado cambios en las credenciales', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar las credenciales: {str(e)}', 'danger')
    
    return redirect(url_for('tasks.view_location', id=location_id))

# API para obtener credenciales del portal
@tasks_bp.route('/locations/<int:location_id>/get-credentials', methods=['POST'])
@login_required
@manager_required
def get_portal_credentials(location_id):
    """Obtiene las credenciales fijas del portal mediante AJAX de forma segura"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or current_user.company_id != location.company_id):
        return jsonify({'success': False, 'error': 'No tienes permiso para obtener las credenciales de este local'}), 403
    
    # Registrar acceso a las credenciales en los logs
    log_activity(f'Acceso a credenciales de portal para local: {location.name}', user_id=current_user.id)
    
    # Obtener el nombre de usuario
    username = location.portal_fixed_username
    
    # Para la contraseña, tenemos un caso especial
    password = location.portal_fixed_password
    
    # Si la contraseña es None, significa que hay una contraseña personalizada (hash)
    # y no podemos mostrarla directamente - mostramos un marcador
    has_custom_password = location.portal_password_hash is not None
    password_placeholder = "********" if has_custom_password else password
    
    return jsonify({
        'success': True,
        'username': username,
        'password': password_placeholder,
        'has_custom_password': has_custom_password
    })

# API para estadísticas en tiempo real
@tasks_bp.route('/api/task-stats')
@login_required
def task_stats():
    """API para obtener estadísticas de tareas"""
    if current_user.is_admin():
        # Administradores ven estadísticas globales
        locations = Location.query.all()
        location_ids = [loc.id for loc in locations]
    elif current_user.is_gerente() and current_user.company_id:
        # Gerentes ven solo su empresa
        locations = Location.query.filter_by(company_id=current_user.company_id).all()
        location_ids = [loc.id for loc in locations]
    else:
        # Otros usuarios no ven nada
        return jsonify({'error': 'No autorizado'}), 403
    
    # Si no hay locales, devolver valores vacíos
    if not location_ids:
        return jsonify({
            'total_tasks': 0,
            'pending_tasks': 0,
            'completed_tasks': 0,
            'completion_rate': 0,
            'tasks_by_priority': {},
            'tasks_by_location': {}
        })
    
    # Contar tareas
    total_tasks = Task.query.filter(Task.location_id.in_(location_ids)).count()
    pending_tasks = Task.query.filter(Task.location_id.in_(location_ids), 
                                     Task.status==TaskStatus.PENDIENTE).count()
    completed_tasks = Task.query.filter(Task.location_id.in_(location_ids), 
                                       Task.status==TaskStatus.COMPLETADA).count()
    
    # Calcular tasa de completado
    completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
    
    # Tareas por prioridad
    tasks_by_priority_query = db.session.query(
        Task.priority,
        db.func.count(Task.id)
    ).filter(
        Task.location_id.in_(location_ids)
    ).group_by(Task.priority).all()
    
    tasks_by_priority = {}
    for priority, count in tasks_by_priority_query:
        if priority:
            tasks_by_priority[priority.value] = count
    
    # Tareas por local
    tasks_by_location_query = db.session.query(
        Location.name,
        db.func.count(Task.id)
    ).join(
        Task, Location.id == Task.location_id
    ).filter(
        Location.id.in_(location_ids)
    ).group_by(Location.name).all()
    
    tasks_by_location = {}
    for location_name, count in tasks_by_location_query:
        tasks_by_location[location_name] = count
    
    return jsonify({
        'total_tasks': total_tasks,
        'pending_tasks': pending_tasks,
        'completed_tasks': completed_tasks,
        'completion_rate': completion_rate,
        'tasks_by_priority': tasks_by_priority,
        'tasks_by_location': tasks_by_location
    })

# Página de etiquetas para usuario local
@tasks_bp.route('/local-user/labels')
@local_user_required
def local_user_labels():
    """Generador de etiquetas para productos - Lista de productos disponibles"""
    user_id = session['local_user_id']
    user = LocalUser.query.get_or_404(user_id)
    location = user.location
    
    # Obtener los productos disponibles para este local
    products = Product.query.filter_by(location_id=location.id, is_active=True).order_by(Product.name).all()
    
    # Obtener impresoras configuradas
    printers = NetworkPrinter.query.filter_by(location_id=location.id).all()
    
    # Filtro de búsqueda (si existe)
    search_query = request.args.get('q', '')
    if search_query:
        products = [p for p in products if search_query.lower() in p.name.lower()]
    
    return render_template('tasks/local_user_labels.html', 
                           title='Selección de Producto',
                           user=user, 
                           location=location, 
                           products=products, 
                           printers=printers,
                           search_query=search_query)


@tasks_bp.route('/local-user/raspberry-pi-setup')
@local_user_required
def raspberry_pi_setup():
    """Página con instrucciones para configurar Raspberry Pi como servidor de impresión"""
    user_id = session['local_user_id']
    user = LocalUser.query.get_or_404(user_id)
    location = user.location
    
    return render_template('tasks/raspberry_pi_setup.html', user=user, location=location)

@tasks_bp.route('/local-user/raspberry-pi-setup-optimized')
@local_user_required
def raspberry_pi_setup_optimized():
    """Página con instrucciones concisas para configurar Raspberry Pi como servidor de impresión"""
    user_id = session['local_user_id']
    user = LocalUser.query.get_or_404(user_id)
    location = user.location
    
    return render_template('tasks/raspberry_pi_setup_optimized.html', user=user, location=location)

# Página de selección de conservación para un producto específico
@tasks_bp.route('/local-user/labels/<int:product_id>')
@local_user_required
def product_conservation_selection(product_id):
    """Selección de tipo de conservación para un producto específico"""
    user_id = session['local_user_id']
    user = LocalUser.query.get_or_404(user_id)
    location = user.location
    
    # Obtener el producto
    product = Product.query.get_or_404(product_id)
    
    # Verificar que el producto pertenece al local del usuario
    if product.location_id != user.location_id:
        flash("El producto seleccionado no pertenece a este local", "danger")
        return redirect(url_for('tasks.local_user_labels'))
    
    # Obtener la fecha y hora actual
    now = datetime.now()
    
    return render_template('tasks/product_conservation_selection.html',
                          title=f'Etiqueta para: {product.name}',
                          user=user,
                          location=location,
                          product=product,
                          now=now)

# Rutas para gestión de impresoras de red
@tasks_bp.route('/dashboard/printers')
@tasks_bp.route('/dashboard/printers/<int:location_id>')
@login_required
@manager_required
def list_printers(location_id=None):
    """Lista de impresoras de red disponibles, filtradas por ubicación si se especifica"""
    if current_user.is_admin():
        # Si no se especifica ubicación, mostrar todas las impresoras para administradores
        if location_id is None:
            printers = NetworkPrinter.query.all()
            locations = Location.query.all()
        else:
            location = Location.query.get_or_404(location_id)
            printers = NetworkPrinter.query.filter_by(location_id=location_id).all()
            locations = [location]
    elif current_user.is_gerente():
        # Los gerentes solo ven las impresoras de sus empresas
        company_ids = [c.id for c in current_user.companies]
        locations = Location.query.filter(Location.company_id.in_(company_ids)).all()
        
        if location_id is None and locations:
            # Si no se especifica ubicación pero hay ubicaciones disponibles, usar la primera
            location_id = locations[0].id
            printers = NetworkPrinter.query.filter_by(location_id=location_id).all()
        elif location_id is not None:
            # Verificar que la ubicación pertenece a una empresa del gerente
            location = Location.query.get_or_404(location_id)
            if location.company_id not in company_ids:
                flash('No tienes permiso para ver las impresoras de esa ubicación.', 'danger')
                return redirect(url_for('tasks.index'))
            printers = NetworkPrinter.query.filter_by(location_id=location_id).all()
        else:
            printers = []
    else:
        # Otros usuarios no tendrían acceso
        printers = []
        locations = []
    
    return render_template('tasks/printer_list.html',
                          title='Gestión de Impresoras',
                          printers=printers,
                          locations=locations,
                          current_location_id=location_id)

@tasks_bp.route('/dashboard/printers/create/<int:location_id>', methods=['GET', 'POST'])
@login_required
@manager_required
def create_printer(location_id):
    """Crear una nueva impresora para una ubicación"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para crear impresoras en esta ubicación.', 'danger')
        return redirect(url_for('tasks.list_printers'))
    
    form = NetworkPrinterForm()
    form.location_id.choices = [(location.id, location.name)]
    form.location_id.data = location.id
    
    if form.validate_on_submit():
        # Crear la instancia de la impresora con todos los campos
        printer = NetworkPrinter(
            name=form.name.data,
            ip_address=form.ip_address.data,
            model=form.model.data,
            port=form.port.data if form.port.data else None,
            api_path=form.api_path.data,
            requires_auth=form.requires_auth.data,
            username=form.username.data if form.requires_auth.data else None,
            password=form.password.data if form.requires_auth.data else None,
            is_default=form.is_default.data,
            location_id=form.location_id.data,
            is_active=True,
            printer_type=form.printer_type.data,
            usb_port=form.usb_port.data if form.printer_type.data == 'raspberry_pi' else None
        )
        
        # Si es la impresora predeterminada, desmarcar otras como predeterminadas
        if printer.is_default:
            other_default_printers = NetworkPrinter.query.filter_by(
                location_id=printer.location_id,
                is_default=True
            ).all()
            
            for other_printer in other_default_printers:
                other_printer.is_default = False
        
        db.session.add(printer)
        db.session.commit()
        
        # Verificar conexión con la impresora
        status = printer.check_status()
        db.session.commit()  # Guardar el estado
        
        log_activity(f'Impresora creada: {printer.name} para la ubicación {location.name}')
        flash(f'Impresora "{printer.name}" creada correctamente. Estado: {"Online" if status else "Offline"}', 'success')
        return redirect(url_for('tasks.list_printers', location_id=location_id))
    
    return render_template('tasks/printer_form.html',
                          title='Crear Nueva Impresora',
                          form=form,
                          location=location)

@tasks_bp.route('/dashboard/printers/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@manager_required
def edit_printer(id):
    """Editar una impresora existente"""
    printer = NetworkPrinter.query.get_or_404(id)
    location = printer.location
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para editar esta impresora.', 'danger')
        return redirect(url_for('tasks.list_printers'))
    
    form = NetworkPrinterForm(obj=printer)
    form.location_id.choices = [(location.id, location.name)]
    
    if form.validate_on_submit():
        printer.name = form.name.data
        printer.ip_address = form.ip_address.data
        printer.model = form.model.data
        # Si se deja vacío el campo del puerto, establecer None
        printer.port = form.port.data if form.port.data else None
        printer.api_path = form.api_path.data
        printer.requires_auth = form.requires_auth.data
        
        # Actualizar los campos de tipo de impresora y puerto USB
        printer.printer_type = form.printer_type.data
        printer.usb_port = form.usb_port.data if form.printer_type.data == 'raspberry_pi' else None
        
        # Solo actualizar credenciales si se requiere autenticación
        if printer.requires_auth:
            printer.username = form.username.data
            if form.password.data:  # Solo actualizar la contraseña si se proporciona una nueva
                printer.password = form.password.data
        else:
            printer.username = None
            printer.password = None
        
        old_default = printer.is_default
        new_default = form.is_default.data
        printer.is_default = new_default
        
        # Si ahora es predeterminada y antes no lo era, desmarcar otras predeterminadas
        if new_default and not old_default:
            other_default_printers = NetworkPrinter.query.filter(
                NetworkPrinter.location_id == printer.location_id,
                NetworkPrinter.id != printer.id,
                NetworkPrinter.is_default == True
            ).all()
            
            for other_printer in other_default_printers:
                other_printer.is_default = False
        
        db.session.commit()
        
        # Verificar conexión con la impresora
        status = printer.check_status()
        db.session.commit()  # Guardar el estado
        
        log_activity(f'Impresora actualizada: {printer.name}')
        flash(f'Impresora "{printer.name}" actualizada correctamente. Estado: {"Online" if status else "Offline"}', 'success')
        return redirect(url_for('tasks.list_printers', location_id=location.id))
    
    return render_template('tasks/printer_form.html',
                          title=f'Editar Impresora: {printer.name}',
                          form=form,
                          printer=printer)

@tasks_bp.route('/dashboard/printers/delete/<int:id>', methods=['POST'])
@login_required
@manager_required
def delete_printer(id):
    """Eliminar una impresora"""
    printer = NetworkPrinter.query.get_or_404(id)
    location = printer.location
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para eliminar esta impresora.', 'danger')
        return redirect(url_for('tasks.list_printers'))
    
    name = printer.name
    location_id = location.id
    db.session.delete(printer)
    db.session.commit()
    
    log_activity(f'Impresora eliminada: {name}')
    flash(f'Impresora "{name}" eliminada correctamente.', 'success')
    return redirect(url_for('tasks.list_printers', location_id=location_id))

@tasks_bp.route('/dashboard/printers/test/<int:id>', methods=['POST'])
@login_required
@manager_required
def test_printer(id):
    """Probar conexión con una impresora"""
    printer = NetworkPrinter.query.get_or_404(id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or printer.location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para probar esta impresora.', 'danger')
        return redirect(url_for('tasks.list_printers'))
    
    status = printer.check_status()
    db.session.commit()  # Guardar el estado
    
    if status:
        flash(f'La impresora "{printer.name}" está en línea.', 'success')
    else:
        flash(f'No se pudo conectar con la impresora "{printer.name}". Por favor, verifica la dirección IP y el puerto.', 'danger')
    
    return redirect(url_for('tasks.list_printers', location_id=printer.location_id))

@tasks_bp.route('/dashboard/printers/set-default/<int:id>', methods=['POST'])
@login_required
@manager_required
def set_default_printer(id):
    """Establecer una impresora como predeterminada"""
    printer = NetworkPrinter.query.get_or_404(id)
    location = printer.location
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para modificar esta impresora.', 'danger')
        return redirect(url_for('tasks.list_printers'))
    
    # Desmarcar todas las impresoras predeterminadas para esta ubicación
    other_default_printers = NetworkPrinter.query.filter(
        NetworkPrinter.location_id == printer.location_id,
        NetworkPrinter.is_default == True
    ).all()
    
    for other_printer in other_default_printers:
        other_printer.is_default = False
    
    # Marcar esta impresora como predeterminada
    printer.is_default = True
    db.session.commit()
    
    log_activity(f'Impresora "{printer.name}" establecida como predeterminada')
    flash(f'Impresora "{printer.name}" establecida como predeterminada.', 'success')
    return redirect(url_for('tasks.list_printers', location_id=printer.location_id))

# API para obtener impresoras (usado por el script brother-printer.js)
@tasks_bp.route('/api/printers/<int:location_id>')
@login_required
def api_get_printers(location_id):
    """API para obtener la lista de impresoras disponibles para una ubicación
    
    Parámetros Query:
    - type: Filtrar por tipo de impresora (direct_network, raspberry_pi)
    """
    # Verificar acceso
    if not current_user.is_admin() and current_user.is_gerente():
        company_ids = [c.id for c in current_user.companies]
        location = Location.query.get_or_404(location_id)
        
        if location.company_id not in company_ids:
            return jsonify({'error': 'Acceso denegado'}), 403
    
    # Obtener tipo de impresora del query string si existe
    printer_type = request.args.get('type', None)
    
    # Obtener impresoras activas
    query = NetworkPrinter.query.filter_by(location_id=location_id, is_active=True)
    
    # Filtrar por tipo si se especifica
    if printer_type:
        query = query.filter_by(printer_type=printer_type)
    
    printers = query.all()
    
    # Formatear respuesta
    printer_list = []
    for printer in printers:
        printer_data = {
            'id': printer.id,
            'name': printer.name,
            'ip_address': printer.ip_address,
            'port': printer.port,
            'model': printer.model,
            'api_path': printer.api_path,
            'requires_auth': printer.requires_auth,
            'username': printer.username if printer.requires_auth else None,
            'password': printer.password if printer.requires_auth else None,
            'is_default': printer.is_default,
            'last_status': printer.last_status,
            'last_status_check': printer.last_status_check.isoformat() if printer.last_status_check else None
        }
        
        # Añadir campos específicos para cada tipo de impresora
        printer_data['printer_type'] = printer.printer_type.value
        
        # Añadir el puerto USB si es una impresora Raspberry Pi
        if printer.printer_type.value == 'raspberry_pi':
            printer_data['usb_port'] = printer.usb_port
        
        printer_list.append(printer_data)
    
    return jsonify({
        'success': True,
        'location_id': location_id,
        'printers': printer_list
    })


@tasks_bp.route('/api/printer/<int:printer_id>')
@login_required
def api_get_printer(printer_id):
    """API para obtener información detallada de una impresora específica"""
    # Obtener la impresora
    printer = NetworkPrinter.query.get_or_404(printer_id)
    
    # Verificar acceso
    if not current_user.is_admin() and current_user.is_gerente():
        company_ids = [c.id for c in current_user.companies]
        location = printer.location
        
        if location.company_id not in company_ids:
            return jsonify({'error': 'Acceso denegado'}), 403
    
    # Formatear respuesta
    printer_data = {
        'id': printer.id,
        'name': printer.name,
        'ip_address': printer.ip_address,
        'port': printer.port,
        'model': printer.model,
        'api_path': printer.api_path,
        'requires_auth': printer.requires_auth,
        'username': printer.username if printer.requires_auth else None,
        'password': printer.password if printer.requires_auth else None,
        'is_default': printer.is_default,
        'last_status': printer.last_status,
        'last_status_check': printer.last_status_check.isoformat() if printer.last_status_check else None,
        'location_id': printer.location_id
    }
    
    # Añadir campos específicos para cada tipo de impresora
    printer_data['printer_type'] = printer.printer_type.value
    
    # Añadir el puerto USB si es una impresora Raspberry Pi
    if printer.printer_type.value == 'raspberry_pi':
        printer_data['usb_port'] = printer.usb_port
    
    return jsonify({
        'success': True,
        'printer': printer_data
    })

# Gestor de etiquetas en la página de tareas
@tasks_bp.route('/dashboard/labels')
@tasks_bp.route('/dashboard/labels/<int:location_id>')
@login_required
@manager_required
def manage_labels(location_id=None):
    """Gestor de etiquetas para la página de tareas, filtrado por ubicación si se especifica"""
    companies = []
    
    # Filtrar empresas según el rol del usuario
    if current_user.is_admin():
        companies = Company.query.all()
    else:
        companies = current_user.companies
    
    # Obtener ubicaciones asociadas a las empresas que puede ver
    company_ids = [c.id for c in companies]
    
    # Si se especifica una ubicación, verificar permisos
    location = None
    if location_id:
        location = Location.query.get_or_404(location_id)
        if location.company_id not in company_ids and not current_user.is_admin():
            flash('No tiene permisos para acceder a esta ubicación', 'danger')
            return redirect(url_for('tasks.index'))
        
        # Filtrar sólo por la ubicación especificada
        locations = [location]
        location_ids = [location_id]
    else:
        # Sin filtro de ubicación, mostrar todas las ubicaciones permitidas
        locations = Location.query.filter(Location.company_id.in_(company_ids)).all()
        location_ids = [loc.id for loc in locations]
    
    # Obtener todos los productos de las ubicaciones permitidas
    products = []
    if locations:
        products = Product.query.filter(Product.location_id.in_(location_ids)).order_by(Product.name).all()
    
    # Obtener etiquetas generadas recientemente (últimos 30 días)
    recent_labels = []
    if products:  # Si hay productos
        thirty_days_ago = datetime.now() - timedelta(days=30)
        
        recent_labels = db.session.query(
            ProductLabel, Product, LocalUser
        ).join(
            Product, ProductLabel.product_id == Product.id
        ).join(
            LocalUser, ProductLabel.local_user_id == LocalUser.id
        ).filter(
            Product.location_id.in_(location_ids),
            ProductLabel.created_at > thirty_days_ago
        ).order_by(
            ProductLabel.created_at.desc()
        ).limit(50).all()
    
    título = f'Etiquetas de {location.name}' if location else 'Gestor de Etiquetas'
    
    return render_template('tasks/manage_labels.html',
                          title=título,
                          companies=companies,
                          locations=locations,
                          selected_location=location,
                          products=products,
                          recent_labels=recent_labels)

@tasks_bp.route('/locations/<int:location_id>/label-editor', methods=['GET', 'POST'])
@login_required
@manager_required
def label_editor(location_id):
    """Editor de diseño de etiquetas para un local"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos
    if not current_user.is_admin() and (not current_user.is_gerente() or location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para editar etiquetas en este local.', 'danger')
        return redirect(url_for('tasks.manage_labels'))
    
    # Buscar plantilla existente o crear una nueva
    template = LabelTemplate.query.filter_by(location_id=location_id, is_default=True).first()
    
    if not template:
        template = LabelTemplate(
            name="Diseño Predeterminado",
            location_id=location_id,
            is_default=True
        )
        db.session.add(template)
        db.session.commit()
    
    form = LabelEditorForm(obj=template)
    
    if form.validate_on_submit():
        # Actualizar la plantilla con los datos del formulario
        form.populate_obj(template)
        template.name = form.layout_name.data
        template.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        log_activity(f'Diseño de etiquetas actualizado para {location.name}')
        flash('Diseño de etiquetas actualizado correctamente.', 'success')
        return redirect(url_for('tasks.manage_labels', location_id=location_id))
    
    return render_template('tasks/label_editor.html',
                          title=f'Editor de Etiquetas - {location.name}',
                          form=form,
                          location=location,
                          template=template)

@tasks_bp.route('/locations/<int:location_id>/label-templates', methods=['GET'])
@login_required
@manager_required
def list_label_templates(location_id):
    """Lista de plantillas de etiquetas para un local"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos
    if not current_user.is_admin() and (not current_user.is_gerente() or location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para ver plantillas de etiquetas en este local.', 'danger')
        return redirect(url_for('tasks.manage_labels'))
    
    templates = LabelTemplate.query.filter_by(location_id=location_id).all()
    
    return render_template('tasks/label_template_list.html',
                          title=f'Plantillas de Etiquetas - {location.name}',
                          location=location,
                          templates=templates)

@tasks_bp.route('/locations/<int:location_id>/label-templates/create', methods=['GET', 'POST'])
@login_required
@manager_required
def create_label_template(location_id):
    """Crear una nueva plantilla de etiquetas"""
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos
    if not current_user.is_admin() and (not current_user.is_gerente() or location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para crear plantillas de etiquetas en este local.', 'danger')
        return redirect(url_for('tasks.manage_labels'))
    
    form = LabelEditorForm()
    
    if form.validate_on_submit():
        template = LabelTemplate(
            name=form.layout_name.data,
            location_id=location_id,
            is_default=False,
            titulo_x=form.titulo_x.data,
            titulo_y=form.titulo_y.data,
            titulo_size=form.titulo_size.data,
            titulo_bold=form.titulo_bold.data,
            conservacion_x=form.conservacion_x.data,
            conservacion_y=form.conservacion_y.data,
            conservacion_size=form.conservacion_size.data,
            conservacion_bold=form.conservacion_bold.data,
            preparador_x=form.preparador_x.data,
            preparador_y=form.preparador_y.data,
            preparador_size=form.preparador_size.data,
            preparador_bold=form.preparador_bold.data,
            fecha_x=form.fecha_x.data,
            fecha_y=form.fecha_y.data,
            fecha_size=form.fecha_size.data,
            fecha_bold=form.fecha_bold.data,
            caducidad_x=form.caducidad_x.data,
            caducidad_y=form.caducidad_y.data,
            caducidad_size=form.caducidad_size.data,
            caducidad_bold=form.caducidad_bold.data,
            caducidad2_x=form.caducidad2_x.data,
            caducidad2_y=form.caducidad2_y.data,
            caducidad2_size=form.caducidad2_size.data,
            caducidad2_bold=form.caducidad2_bold.data
        )
        
        db.session.add(template)
        db.session.commit()
        
        log_activity(f'Plantilla de etiquetas creada: {template.name} para {location.name}')
        flash(f'Plantilla "{template.name}" creada correctamente.', 'success')
        return redirect(url_for('tasks.list_label_templates', location_id=location_id))
    
    return render_template('tasks/label_editor.html',
                          title=f'Nueva Plantilla de Etiquetas - {location.name}',
                          form=form,
                          location=location,
                          is_new=True)

@tasks_bp.route('/label-templates/<int:template_id>/edit', methods=['GET', 'POST'])
@login_required
@manager_required
def edit_label_template(template_id):
    """Editar una plantilla de etiquetas existente"""
    template = LabelTemplate.query.get_or_404(template_id)
    location = Location.query.get_or_404(template.location_id)
    
    # Verificar permisos
    if not current_user.is_admin() and (not current_user.is_gerente() or location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para editar plantillas de etiquetas en este local.', 'danger')
        return redirect(url_for('tasks.manage_labels'))
    
    form = LabelEditorForm(obj=template)
    
    if form.validate_on_submit():
        # Actualizar la plantilla con los datos del formulario
        form.populate_obj(template)
        template.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        log_activity(f'Plantilla de etiquetas actualizada: {template.name}')
        flash(f'Plantilla "{template.name}" actualizada correctamente.', 'success')
        return redirect(url_for('tasks.list_label_templates', location_id=location.id))
    
    return render_template('tasks/label_editor.html',
                          title=f'Editar Plantilla: {template.name}',
                          form=form,
                          location=location,
                          template=template)

@tasks_bp.route('/label-templates/<int:template_id>/delete', methods=['POST'])
@login_required
@manager_required
def delete_label_template(template_id):
    """Eliminar una plantilla de etiquetas"""
    template = LabelTemplate.query.get_or_404(template_id)
    location = Location.query.get_or_404(template.location_id)
    
    # Verificar permisos
    if not current_user.is_admin() and (not current_user.is_gerente() or location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para eliminar plantillas de etiquetas en este local.', 'danger')
        return redirect(url_for('tasks.manage_labels'))
    
    # No permitir eliminar la plantilla predeterminada
    if template.is_default:
        flash('No puedes eliminar la plantilla predeterminada.', 'warning')
        return redirect(url_for('tasks.list_label_templates', location_id=location.id))
    
    name = template.name
    location_id = location.id
    
    db.session.delete(template)
    db.session.commit()
    
    log_activity(f'Plantilla de etiquetas eliminada: {name}')
    flash(f'Plantilla "{name}" eliminada correctamente.', 'success')
    return redirect(url_for('tasks.list_label_templates', location_id=location_id))

@tasks_bp.route('/label-templates/<int:template_id>/set-default', methods=['POST'])
@login_required
@manager_required
def set_default_label_template(template_id):
    """Establecer una plantilla como predeterminada"""
    template = LabelTemplate.query.get_or_404(template_id)
    location = Location.query.get_or_404(template.location_id)
    
    # Verificar permisos
    if not current_user.is_admin() and (not current_user.is_gerente() or location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para modificar plantillas de etiquetas en este local.', 'danger')
        return redirect(url_for('tasks.manage_labels'))
    
    # Quitar el estado predeterminado de todas las plantillas de esta ubicación
    default_templates = LabelTemplate.query.filter_by(location_id=location.id, is_default=True).all()
    for default_template in default_templates:
        default_template.is_default = False
    
    # Establecer esta plantilla como predeterminada
    template.is_default = True
    db.session.commit()
    
    log_activity(f'Plantilla de etiquetas establecida como predeterminada: {template.name}')
    flash(f'Plantilla "{template.name}" establecida como predeterminada.', 'success')
    return redirect(url_for('tasks.list_label_templates', location_id=location.id))

@tasks_bp.route('/dashboard/labels/download-template')
@login_required
@manager_required
def download_excel_template():
    """Descargar plantilla vacía en Excel para importación de productos"""
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Añadir encabezados
    ws['A1'] = "Nombre"
    ws['B1'] = "Descripción"
    ws['C1'] = "Vida útil (días)"
    ws['D1'] = "Descongelación (horas)"
    ws['E1'] = "Refrigeración (horas)"
    ws['F1'] = "Gastro (horas)"
    ws['G1'] = "Caliente (horas)"
    ws['H1'] = "Seco (horas)"
    
    # Añadir una fila de ejemplo
    ws['A2'] = "Ejemplo Producto"
    ws['B2'] = "Descripción de ejemplo"
    ws['C2'] = 12  # Vida útil secundaria en días
    ws['D2'] = 48  # Horas para descongelación (2 días)
    ws['E2'] = 72  # Horas para refrigeración (3 días) 
    ws['F2'] = 96  # Horas para gastro (4 días)
    ws['G2'] = 2   # Horas para caliente
    ws['H2'] = 168 # Horas para seco (7 días)
    
    # Guardar a un objeto BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear nombre de archivo para la plantilla
    filename = f"plantilla_productos_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@tasks_bp.route('/dashboard/labels/export/<int:location_id>')
@login_required
@manager_required
def export_labels_excel(location_id):
    """Exportar lista de productos y tipos de conservación a Excel"""
    # Verificar que el local existe y el usuario tiene permisos
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or 
                                       location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para exportar datos de este local.', 'danger')
        return redirect(url_for('tasks.manage_labels'))
    
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Añadir encabezados
    ws['A1'] = "Nombre"
    ws['B1'] = "Descripción"
    ws['C1'] = "Vida útil (días)"
    ws['D1'] = "Descongelación (horas)"
    ws['E1'] = "Refrigeración (horas)"
    ws['F1'] = "Gastro (horas)"
    ws['G1'] = "Caliente (horas)"
    ws['H1'] = "Seco (horas)"
    
    # Obtener productos de este local
    products = Product.query.filter_by(location_id=location_id).order_by(Product.name).all()
    
    # Rellenar datos
    row = 2
    for product in products:
        ws[f'A{row}'] = product.name
        ws[f'B{row}'] = product.description or ""
        ws[f'C{row}'] = product.shelf_life_days
        
        # Buscar horas de conservación para cada tipo
        for conservation in product.conservation_types:
            # Usar horas directamente
            hours_valid = conservation.hours_valid
            
            if conservation.conservation_type == ConservationType.DESCONGELACION:
                ws[f'D{row}'] = hours_valid
            elif conservation.conservation_type == ConservationType.REFRIGERACION:
                ws[f'E{row}'] = hours_valid
            elif conservation.conservation_type == ConservationType.GASTRO:
                ws[f'F{row}'] = hours_valid
            elif conservation.conservation_type == ConservationType.CALIENTE:
                ws[f'G{row}'] = hours_valid
            elif conservation.conservation_type == ConservationType.SECO:
                ws[f'H{row}'] = hours_valid
        
        row += 1
    
    # Guardar a un objeto BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear nombre de archivo basado en la ubicación
    filename = f"productos_{location.name.replace(' ', '_').lower()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@tasks_bp.route('/dashboard/labels/import/<int:location_id>', methods=['POST'])
@login_required
@manager_required
def import_labels_excel(location_id):
    """Importar lista de productos y tipos de conservación desde Excel"""
    # Verificar que el local existe y el usuario tiene permisos
    location = Location.query.get_or_404(location_id)
    
    # Verificar permisos (admin o gerente de la empresa)
    if not current_user.is_admin() and (not current_user.is_gerente() or 
                                       location.company_id not in [c.id for c in current_user.companies]):
        flash('No tienes permiso para importar datos a este local.', 'danger')
        return redirect(url_for('tasks.manage_labels'))
    
    # Comprobar si se ha subido un archivo
    if 'excel_file' not in request.files:
        flash('No se ha seleccionado ningún archivo.', 'danger')
        return redirect(url_for('tasks.manage_labels', location_id=location_id))
        
    file = request.files['excel_file']
    
    if file.filename == '':
        flash('No se ha seleccionado ningún archivo.', 'danger')
        return redirect(url_for('tasks.manage_labels', location_id=location_id))
    
    if not file.filename.endswith('.xlsx'):
        flash('El archivo debe ser un archivo Excel (.xlsx).', 'danger')
        return redirect(url_for('tasks.manage_labels', location_id=location_id))
    
    try:
        # Cargar el archivo Excel
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Contar productos actualizados y creados
        updated = 0
        created = 0
        errors = 0
        
        # Empezar desde la fila 2 (después de encabezados)
        for row in range(2, ws.max_row + 1):
            try:
                # Obtener datos del producto
                product_name = ws[f'A{row}'].value
                product_description = ws[f'B{row}'].value or ""
                
                # Si no hay nombre, ignorar esta fila
                if not product_name:
                    continue
                
                # Vida útil en días (caducidad secundaria)
                shelf_life_days = ws[f'C{row}'].value
                shelf_life_days = int(shelf_life_days) if shelf_life_days is not None else 0
                
                # Horas para cada tipo de conservación
                hours_descongelacion = ws[f'D{row}'].value
                hours_refrigeracion = ws[f'E{row}'].value
                hours_gastro = ws[f'F{row}'].value
                hours_caliente = ws[f'G{row}'].value
                hours_seco = ws[f'H{row}'].value
                
                # Usar horas directamente para almacenar en la base de datos
                hours_descongelacion = int(hours_descongelacion) if hours_descongelacion is not None else None
                hours_refrigeracion = int(hours_refrigeracion) if hours_refrigeracion is not None else None
                hours_gastro = int(hours_gastro) if hours_gastro is not None else None
                hours_caliente = int(hours_caliente) if hours_caliente is not None else None
                hours_seco = int(hours_seco) if hours_seco is not None else None
                
                # Buscar producto existente por nombre en esta ubicación
                product = Product.query.filter_by(name=product_name, location_id=location_id).first()
                
                if product:
                    # Actualizar producto existente
                    product.name = product_name
                    product.description = product_description
                    product.shelf_life_days = shelf_life_days
                    updated += 1
                else:
                    # Crear nuevo producto
                    product = Product(
                        name=product_name,
                        description=product_description,
                        shelf_life_days=shelf_life_days,
                        location_id=location_id
                    )
                    db.session.add(product)
                    db.session.flush()  # Para obtener el ID generado
                    created += 1
                
                # Actualizar tipos de conservación
                conservation_types = [
                    (ConservationType.DESCONGELACION, hours_descongelacion),
                    (ConservationType.REFRIGERACION, hours_refrigeracion),
                    (ConservationType.GASTRO, hours_gastro),
                    (ConservationType.CALIENTE, hours_caliente),
                    (ConservationType.SECO, hours_seco)
                ]
                
                for cons_type, hours in conservation_types:
                    if hours is not None and hours > 0:
                        # Buscar conservación existente o crear una nueva
                        conservation = ProductConservation.query.filter_by(
                            product_id=product.id, 
                            conservation_type=cons_type
                        ).first()
                        
                        if conservation:
                            conservation.hours_valid = hours
                        else:
                            conservation = ProductConservation(
                                product_id=product.id,
                                conservation_type=cons_type,
                                hours_valid=hours
                            )
                            db.session.add(conservation)
            
            except Exception as e:
                # Registrar error y continuar con la siguiente fila
                errors += 1
                current_app.logger.error(f"Error al importar fila {row}: {str(e)}")
        
        # Guardar todos los cambios
        db.session.commit()
        
        # Mostrar mensaje de éxito
        if errors > 0:
            flash(f'Importación completada con {created} productos creados, {updated} actualizados y {errors} errores.', 'warning')
        else:
            flash(f'Importación completada con éxito: {created} productos creados y {updated} actualizados.', 'success')
        
        # Registrar actividad
        log_activity(f'Importación de productos desde Excel para {location.name}: {created} creados, {updated} actualizados')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')
        current_app.logger.error(f"Error al importar Excel: {str(e)}")
    
    return redirect(url_for('tasks.manage_labels', location_id=location_id))

@tasks_bp.route('/local-user/generate-labels', methods=['POST'])
@local_user_required
def generate_labels():
    """Endpoint simplificado para generar e imprimir etiquetas directamente"""
    try:
        # Obtener el usuario local
        user_id = session['local_user_id']
        user = LocalUser.query.get_or_404(user_id)
        
        # Obtener datos del formulario
        product_id = request.form.get('product_id', type=int)
        conservation_type_str = request.form.get('conservation_type')
        quantity = request.form.get('quantity', type=int, default=1)
        
        # Validaciones básicas
        if not product_id or not conservation_type_str:
            return "Error: Faltan datos obligatorios", 400
            
        # Validar cantidad (entre 1 y 100)
        quantity = max(1, min(100, quantity))
        
        # Obtener el producto
        product = Product.query.get_or_404(product_id)
        
        # Verificar que el producto pertenece al local del usuario
        if product.location_id != user.location_id:
            return "Error: El producto no pertenece a este local", 403
        
        # Convertir string a enum
        conservation_type = None
        for ct in ConservationType:
            if ct.value == conservation_type_str:
                conservation_type = ct
                break
                
        if not conservation_type:
            return "Error: Tipo de conservación no válido", 400
        
        # Obtener configuración de conservación específica
        conservation = ProductConservation.query.filter_by(
            product_id=product.id, 
            conservation_type=conservation_type
        ).first()
        
        # Fecha y hora actual
        now = datetime.now()
        
        # Calcular fecha de caducidad primaria (por conservación)
        if conservation:
            # Usar la configuración específica del producto
            expiry_datetime = now + timedelta(hours=conservation.hours_valid)
        else:
            # Valores predeterminados por tipo
            hours_map = {
                ConservationType.DESCONGELACION: 24,  # 1 día
                ConservationType.REFRIGERACION: 72,   # 3 días
                ConservationType.GASTRO: 48,          # 2 días
                ConservationType.CALIENTE: 2,         # 2 horas
                ConservationType.SECO: 720            # 30 días
            }
            hours = hours_map.get(conservation_type, 24)  # 24h por defecto
            expiry_datetime = now + timedelta(hours=hours)
            
        # Calcular fecha de caducidad secundaria (por vida útil en días)
        secondary_expiry_date = None
        if product.shelf_life_days > 0:
            secondary_expiry_date = product.get_shelf_life_expiry(now)
        
        # Obtener la plantilla de etiquetas predeterminada para este local
        template = LabelTemplate.query.filter_by(location_id=user.location_id, is_default=True).first()
        
        # Si no existe una plantilla predeterminada, crear una
        if not template:
            template = LabelTemplate(
                name="Diseño Predeterminado",
                location_id=user.location_id,
                is_default=True
            )
            db.session.add(template)
            db.session.commit()
        
        # Registrar las etiquetas en la base de datos
        try:
            for _ in range(quantity):
                label = ProductLabel(
                    product_id=product.id,
                    local_user_id=user.id,
                    conservation_type=conservation_type,
                    expiry_date=expiry_datetime.date()
                )
                db.session.add(label)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error al registrar etiquetas: {str(e)}")
            # Continuar generando etiquetas aunque falle el registro
        
        # Verificar si necesitamos generar automáticamente una etiqueta de refrigeración después de descongelación
        auto_generate_refrigeration = False
        refrigeration_conservation_type = None
        refrigeration_expiry_datetime = None
        
        if conservation_type == ConservationType.DESCONGELACION:
            auto_generate_refrigeration = True
            
            # Obtener el tipo de conservación de refrigeración
            for ct in ConservationType:
                if ct.value == "refrigeracion":
                    refrigeration_conservation_type = ct
                    break
            
            # Obtener la configuración de refrigeración (si existe)
            ref_conservation = None
            if refrigeration_conservation_type:
                ref_conservation = ProductConservation.query.filter_by(
                    product_id=product.id, 
                    conservation_type=refrigeration_conservation_type
                ).first()
            
            # Calcular fecha de caducidad para refrigeración a partir de la fecha de caducidad de descongelación
            if ref_conservation:
                # La hora de inicio de refrigeración es la hora de finalización de descongelación
                refrigeration_expiry_datetime = expiry_datetime + timedelta(hours=ref_conservation.hours_valid)
            else:
                # Usar valor predeterminado si no hay configuración específica (3 días)
                refrigeration_expiry_datetime = expiry_datetime + timedelta(hours=72)
                
            # Registrar etiquetas de refrigeración
            try:
                for _ in range(quantity):
                    label = ProductLabel(
                        product_id=product.id,
                        local_user_id=user.id,
                        conservation_type=refrigeration_conservation_type,
                        expiry_date=refrigeration_expiry_datetime.date()
                    )
                    db.session.add(label)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error al registrar etiquetas de refrigeración: {str(e)}")
                # Continuar de todos modos
        
        # Generar HTML para impresión directa, incluyendo las etiquetas de refrigeración si corresponde
        return render_template(
            'tasks/print_labels.html',
            product=product,
            user=user,
            location=user.location,
            conservation_type=conservation_type,
            now=now,
            expiry_datetime=expiry_datetime,
            secondary_expiry_date=secondary_expiry_date,
            quantity=quantity,
            template=template,
            auto_generate_refrigeration=auto_generate_refrigeration,
            refrigeration_conservation_type=refrigeration_conservation_type,
            refrigeration_expiry_datetime=refrigeration_expiry_datetime,
            refrigeration_start_time=expiry_datetime  # La hora de inicio de refrigeración es la hora de fin de descongelación
        )
        
    except Exception as e:
        current_app.logger.error(f"Error en generate_labels: {str(e)}")
        return "Error al generar etiquetas. Inténtelo nuevamente.", 500

@tasks_bp.route('/admin/products')
@tasks_bp.route('/admin/products/<int:location_id>')
@login_required
@manager_required
def list_products(location_id=None):
    """Lista de productos, filtrada por ubicación si se especifica"""
    companies = []
    
    # Filtrar empresas según el rol del usuario
    if current_user.is_admin():
        companies = Company.query.all()
    else:
        companies = current_user.companies
    
    # Obtener ubicaciones asociadas a las empresas que puede ver
    company_ids = [c.id for c in companies]
    
    # Si se especifica una ubicación, verificar permisos
    location = None
    if location_id:
        location = Location.query.get_or_404(location_id)
        if location.company_id not in company_ids and not current_user.is_admin():
            flash('No tiene permisos para acceder a esta ubicación', 'danger')
            return redirect(url_for('tasks.list_products'))
        
        # Filtrar sólo por la ubicación especificada
        locations = [location]
        location_ids = [location_id]
    else:
        # Sin filtro de ubicación, mostrar todas las ubicaciones permitidas
        locations = Location.query.filter(Location.company_id.in_(company_ids)).all()
        location_ids = [loc.id for loc in locations]
    
    # Si no hay ubicaciones, redireccionar a crear ubicación
    if not locations:
        flash('Primero debe crear una ubicación', 'warning')
        return redirect(url_for('tasks.list_locations'))
    
    # Obtener productos de esas ubicaciones
    products = Product.query.filter(Product.location_id.in_(location_ids)).order_by(Product.name).all()
    
    título = f'Productos de {location.name}' if location else 'Productos'
    
    return render_template(
        'tasks/product_list.html',
        title=título,
        products=products,
        locations=locations,
        selected_location=location
    )

@tasks_bp.route('/admin/products/create', methods=['GET', 'POST'])
@tasks_bp.route('/admin/products/create/<int:location_id>', methods=['GET', 'POST'])
@login_required
@manager_required
def create_product(location_id=None):
    """Crear nuevo producto, opcionalmente preseleccionando una ubicación"""
    form = ProductForm()
    
    # Si se proporciona un ID de ubicación, verificar permisos
    preselected_location = None
    if location_id:
        preselected_location = Location.query.get_or_404(location_id)
        if not current_user.is_admin():
            company_ids = [c.id for c in current_user.companies]
            if preselected_location.company_id not in company_ids:
                flash('No tiene permisos para crear productos en esta ubicación', 'danger')
                return redirect(url_for('tasks.list_products'))
    
    # Opciones de ubicaciones
    locations = []
    if current_user.is_admin():
        locations = Location.query.order_by(Location.name).all()
    else:
        company_ids = [c.id for c in current_user.companies]
        locations = Location.query.filter(Location.company_id.in_(company_ids)).order_by(Location.name).all()
    
    form.location_id.choices = [(l.id, f"{l.name} ({l.company.name})") for l in locations]
    
    # Si hay una ubicación preseleccionada, establecerla como valor predeterminado
    if preselected_location and request.method == 'GET':
        form.location_id.data = preselected_location.id
    
    if form.validate_on_submit():
        product = Product(
            name=form.name.data,
            description=form.description.data,
            shelf_life_days=form.shelf_life_days.data or 0,
            is_active=form.is_active.data,
            location_id=form.location_id.data
        )
        
        db.session.add(product)
        
        try:
            db.session.commit()
            flash(f'Producto "{product.name}" creado correctamente', 'success')
            return redirect(url_for('tasks.list_products'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear producto: {str(e)}', 'danger')
    
    return render_template(
        'tasks/product_form.html',
        title='Nuevo Producto',
        form=form,
        is_edit=False
    )

@tasks_bp.route('/admin/products/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@manager_required
def edit_product(id):
    """Editar producto existente"""
    product = Product.query.get_or_404(id)
    
    # Verificar permisos
    if not current_user.is_admin():
        company_ids = [c.id for c in current_user.companies]
        location = Location.query.get_or_404(product.location_id)
        if location.company_id not in company_ids:
            flash('No tiene permisos para editar este producto', 'danger')
            return redirect(url_for('tasks.list_products'))
    
    form = ProductForm(obj=product)
    
    # Opciones de ubicaciones
    locations = []
    if current_user.is_admin():
        locations = Location.query.order_by(Location.name).all()
    else:
        company_ids = [c.id for c in current_user.companies]
        locations = Location.query.filter(Location.company_id.in_(company_ids)).order_by(Location.name).all()
    
    form.location_id.choices = [(l.id, f"{l.name} ({l.company.name})") for l in locations]
    
    if form.validate_on_submit():
        product.name = form.name.data
        product.description = form.description.data
        product.shelf_life_days = form.shelf_life_days.data or 0
        product.is_active = form.is_active.data
        product.location_id = form.location_id.data
        
        try:
            db.session.commit()
            flash(f'Producto "{product.name}" actualizado correctamente', 'success')
            return redirect(url_for('tasks.list_products'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar producto: {str(e)}', 'danger')
    
    return render_template(
        'tasks/product_form.html',
        title=f'Editar Producto: {product.name}',
        form=form,
        product=product,
        is_edit=True
    )

@tasks_bp.route('/admin/products/<int:id>/conservations', methods=['GET', 'POST'])
@login_required
@manager_required
def manage_product_conservations(id):
    """Gestionar tipos de conservación para un producto"""
    product = Product.query.get_or_404(id)
    
    # Verificar permisos
    if not current_user.is_admin():
        company_ids = [c.id for c in current_user.companies]
        location = Location.query.get_or_404(product.location_id)
        if location.company_id not in company_ids:
            flash('No tiene permisos para gestionar este producto', 'danger')
            return redirect(url_for('tasks.list_products'))
    
    # Obtener configuraciones de conservación existentes
    conservations = ProductConservation.query.filter_by(product_id=product.id).all()
    
    # Diccionario para facilitar acceso
    conservation_dict = {}
    for conservation in conservations:
        conservation_dict[conservation.conservation_type.value] = conservation
    
    # Si es GET y hay un tipo de conservación seleccionado, prellenar el formulario
    selected_conservation = None
    selected_type = request.args.get('type')
    if request.method == 'GET' and selected_type:
        for cons in conservations:
            if cons.conservation_type.value == selected_type:
                selected_conservation = cons
                break
    
    # Crear formulario con o sin objeto preexistente
    form = ProductConservationForm(obj=selected_conservation)
    
    # Si se proporciona un tipo en la URL, preseleccionarlo
    if request.method == 'GET' and selected_type:
        form.conservation_type.data = selected_type
    
    if form.validate_on_submit():
        conservation_type_str = form.conservation_type.data
        conservation_type = None
        
        # Convertir string a enum
        for ct in ConservationType:
            if ct.value == conservation_type_str:
                conservation_type = ct
                break
        
        if not conservation_type:
            flash('Tipo de conservación no válido', 'danger')
            return redirect(url_for('tasks.manage_product_conservations', id=product.id))
        
        # Buscar si ya existe esta configuración
        conservation = ProductConservation.query.filter_by(
            product_id=product.id,
            conservation_type=conservation_type
        ).first()
        
        # Obtener horas directamente del formulario
        hours_valid = form.hours_valid.data
        
        # Debug logging
        current_app.logger.debug(f"Horas recibidas: {hours_valid}")
        
        if conservation:
            # Actualizar existente
            conservation.hours_valid = hours_valid
        else:
            # Crear nueva
            conservation = ProductConservation(
                product_id=product.id,
                conservation_type=conservation_type,
                hours_valid=hours_valid
            )
            db.session.add(conservation)
        
        try:
            db.session.commit()
            # Verificar que se guardó correctamente
            db.session.refresh(conservation)
            current_app.logger.debug(f"Guardado en BD: {conservation.hours_valid} horas")
            
            flash('Configuración de conservación guardada correctamente', 'success')
            return redirect(url_for('tasks.manage_product_conservations', id=product.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar configuración: {str(e)}', 'danger')
    
    return render_template(
        'tasks/product_conservations.html',
        title=f'Conservación: {product.name}',
        product=product,
        form=form,
        conservations=conservations,
        conservation_dict=conservation_dict
    )

# API para gestión de impresoras desde el portal de usuarios
@tasks_bp.route('/api/printers/location/<int:location_id>')
def api_get_location_printer(location_id):
    """API para obtener las impresoras disponibles para una ubicación"""
    # Verificar si hay algún tipo de autenticación (portal o general)
    if ('portal_authenticated' not in session or not session['portal_authenticated']) and not current_user.is_authenticated:
        current_app.logger.warning(f"Acceso no autorizado a API de impresora para ubicación {location_id}")
        return jsonify({'success': False, 'error': 'unauthorized', 'message': 'No autorizado'}), 403
    
    # Si es usuario autenticado del portal, verificar que tenga acceso a esta ubicación
    if 'portal_authenticated' in session and session['portal_authenticated'] and session.get('location_id') != location_id:
        current_app.logger.warning(f"Usuario de portal intenta acceder a ubicación no permitida: {location_id}")
        return jsonify({'success': False, 'error': 'forbidden', 'message': 'Acceso no permitido a esta ubicación'}), 403
    
    # Buscar todas las impresoras activas para esta ubicación
    printers = NetworkPrinter.query.filter_by(
        location_id=location_id,
        is_active=True
    ).all()
    
    if printers:
        # Crear lista de impresoras para devolver (sin contraseñas por seguridad)
        printer_list = []
        for printer in printers:
            printer_list.append({
                'id': printer.id,
                'name': printer.name,
                'ip_address': printer.ip_address,
                'port': printer.port,
                'model': printer.model,
                'api_path': printer.api_path,
                'printer_type': printer.printer_type,
                'requires_auth': printer.requires_auth,
                'username': printer.username if printer.requires_auth else None,
                'usb_port': printer.usb_port,
                'is_default': printer.is_default,
                'last_status': printer.last_status,
                'last_status_check': printer.last_status_check.isoformat() if printer.last_status_check else None
            })
        
        return jsonify({
            'success': True,
            'printers': printer_list
        })
    else:
        return jsonify({
            'success': False,
            'error': 'no_printer',
            'message': 'No hay impresora configurada para esta ubicación',
            'printers': []
        })

@tasks_bp.route('/api/printers/check/<int:location_id>')
def api_check_printer_status(location_id):
    """API para comprobar el estado de la impresora de una ubicación"""
    # Verificar si hay algún tipo de autenticación (portal o general)
    if ('portal_authenticated' not in session or not session['portal_authenticated']) and not current_user.is_authenticated:
        current_app.logger.warning(f"Acceso no autorizado a API de verificación de impresora para ubicación {location_id}")
        return jsonify({'success': False, 'error': 'unauthorized', 'message': 'No autorizado'}), 403
    
    # Si es usuario autenticado del portal, verificar que tenga acceso a esta ubicación
    if 'portal_authenticated' in session and session['portal_authenticated'] and session.get('location_id') != location_id:
        current_app.logger.warning(f"Usuario de portal intenta verificar impresora de ubicación no permitida: {location_id}")
        return jsonify({'success': False, 'error': 'forbidden', 'message': 'Acceso no permitido a esta ubicación'}), 403
    
    # Buscar la impresora predeterminada para esta ubicación
    printer = NetworkPrinter.query.filter_by(
        location_id=location_id, 
        is_default=True, 
        is_active=True
    ).first()
    
    # Si no hay una predeterminada, obtener cualquier impresora activa
    if not printer:
        printer = NetworkPrinter.query.filter_by(
            location_id=location_id,
            is_active=True
        ).first()
    
    if not printer:
        return jsonify({
            'success': False,
            'error': 'no_printer',
            'message': 'No hay impresora configurada para esta ubicación'
        })
    
    # Comprobar el estado de la impresora
    is_online = printer.check_status()
    
    # Actualizar la base de datos con el nuevo estado
    db.session.commit()
    
    return jsonify({
        'success': True,
        'online': is_online,
        'printer_id': printer.id,
        'printer_name': printer.name,
        'status': printer.last_status,
        'last_check': printer.last_status_check.isoformat() if printer.last_status_check else None
    })

@tasks_bp.route('/api/printers/test-connection', methods=['POST'])
def api_test_printer_connection():
    """API para probar la conexión con una impresora sin guardar la configuración"""
    # Verificar si hay algún tipo de autenticación (portal o general)
    if ('portal_authenticated' not in session or not session['portal_authenticated']) and not current_user.is_authenticated:
        current_app.logger.warning(f"Acceso no autorizado a API de prueba de impresora")
        return jsonify({'success': False, 'error': 'unauthorized', 'message': 'No autorizado'}), 403
    
    # Obtener datos del formulario
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'invalid_request', 'message': 'Datos no válidos'}), 400
    
    ip_address = data.get('ip_address')
    port = data.get('port', 80)
    
    if not ip_address:
        return jsonify({'success': False, 'error': 'missing_data', 'message': 'Dirección IP requerida'}), 400
    
    try:
        # Intentar conectarse al puerto de la impresora
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(2.0)  # Timeout de 2 segundos
        result = s.connect_ex((ip_address, port))
        s.close()
        
        if result == 0:
            return jsonify({
                'success': True,
                'message': 'Conexión exitosa'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'connection_failed',
                'message': f'No se pudo conectar a la impresora en {ip_address}:{port}'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': 'test_error',
            'message': f'Error al probar la conexión: {str(e)}'
        })

@tasks_bp.route('/api/printers/save', methods=['POST'])
def api_save_printer_config():
    """API para guardar la configuración de una impresora"""
    # Verificar si hay algún tipo de autenticación (portal o general)
    if ('portal_authenticated' not in session or not session['portal_authenticated']) and not current_user.is_authenticated:
        current_app.logger.warning(f"Acceso no autorizado a API de guardado de impresora")
        return jsonify({'success': False, 'error': 'unauthorized', 'message': 'No autorizado'}), 403
    
    # Obtener datos del formulario
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'invalid_request', 'message': 'Datos no válidos'}), 400
    
    # Obtener datos necesarios
    location_id = data.get('location_id')
    if not location_id or session.get('location_id') != location_id:
        return jsonify({'success': False, 'error': 'forbidden', 'message': 'Acceso no permitido a esta ubicación'}), 403
    
    name = data.get('name')
    ip_address = data.get('ip_address')
    port = data.get('port', 80)
    model = data.get('model', '')
    api_path = data.get('api_path', '/brother_d/printer/print')
    requires_auth = data.get('requires_auth', False)
    username = data.get('username', '')
    password = data.get('password', '')
    is_default = data.get('is_default', True)
    
    if not name or not ip_address:
        return jsonify({'success': False, 'error': 'missing_data', 'message': 'Nombre e IP son requeridos'}), 400
    
    try:
        # Buscar si ya existe una impresora para esta ubicación
        printer = NetworkPrinter.query.filter_by(
            location_id=location_id,
            is_default=True,
            is_active=True
        ).first()
        
        # Si es una impresora nueva
        if not printer:
            printer = NetworkPrinter(
                name=name,
                ip_address=ip_address,
                port=port if port else None,
                model=model,
                api_path=api_path,
                requires_auth=requires_auth,
                username=username if requires_auth else '',
                password=password if requires_auth and password else '',
                is_default=is_default,
                is_active=True,
                location_id=location_id
            )
            db.session.add(printer)
        else:
            # Actualizar la impresora existente
            printer.name = name
            printer.ip_address = ip_address
            printer.port = port if port else None
            printer.model = model
            printer.api_path = api_path
            printer.requires_auth = requires_auth
            printer.username = username if requires_auth else ''
            # Solo actualizar la contraseña si se proporciona una nueva
            if requires_auth and password:
                printer.password = password
            printer.is_default = is_default
        
        # Si esta impresora es la predeterminada, desmarcar cualquier otra predeterminada
        if is_default:
            other_printers = NetworkPrinter.query.filter(
                NetworkPrinter.location_id == location_id,
                NetworkPrinter.is_default == True,
                NetworkPrinter.id != (printer.id if printer.id else -1)
            ).all()
            
            for other in other_printers:
                other.is_default = False
        
        # Verificar la conexión con la impresora
        printer.check_status()
        
        # Guardar cambios
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Configuración guardada correctamente',
            'printer': {
                'id': printer.id,
                'name': printer.name,
                'ip_address': printer.ip_address,
                'port': printer.port,
                'model': printer.model,
                'api_path': printer.api_path,
                'requires_auth': printer.requires_auth,
                'username': printer.username if printer.requires_auth else None,
                'is_default': printer.is_default,
                'last_status': printer.last_status,
                'last_status_check': printer.last_status_check.isoformat() if printer.last_status_check else None
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': 'save_error',
            'message': f'Error al guardar la configuración: {str(e)}'
        })